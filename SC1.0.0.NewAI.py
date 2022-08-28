import tkinter
from tkinter import *
import customtkinter
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl import load_workbook
import pandas as pd
import os
from pathlib import Path

highlight = NamedStyle(name="highlight")
bd = Side(style='thin', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = GradientFill(stop=("CEF0FA", "CEF0FA"))
highlight.font = Font(name='Tahoma',
                      size=11,
                      bold=True,
                      italic=False,
                      vertAlign=None,
                      underline='none',
                      strike=False,
                      color='FF000000')
fill = PatternFill(fill_type=None,
                   start_color='FFFFFFFF',
                   end_color='FF000000')

highlight.alignment = Alignment(horizontal="center", vertical="center")
highlight.alignment = Alignment(horizontal='center',
                                vertical='center',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0)

customtkinter.set_appearance_mode("dark")  # Modes: "System" (standard), "dark", "light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

app = customtkinter.CTk()
app.geometry("400x580")
app.title("Shifts creator")

wb2 = load_workbook('history.xlsx')
wb = Workbook()
ws = wb.active
wb.add_named_style(highlight)

ws.append(['Sun-Day', 'Sun-Night', 'Mon-Day', 'Mon-Night', 'Tue-Day', 'Tue-Night', 'Wed-Day', 'Wed-Night', 'Thu-Day',
           'Thu-Night', 'Fri-Day', 'Sat-Night'])
flag = 0
val = ""
maxst = ""
maxfr = ""
Alyona_st = 0
Alyona_fr = 0
Alex_st = 0
Alex_fr = 0
Ofir_st = 0
Ofir_fr = 0
Yair_st = 0
Yair_fr = 0
Almog_st = 0
Almog_fr = 0
Ran_st = 0
Ran_fr = 0
Sahar_st = 0
Sahar_fr = 0
temp = " "
SundayM =""
SundayN =""
MondayM =""
MondayN =""
TuesdayM =""
TuesdayN =""
WednesdayM =""
WednesdayN =""
ThursdayM =""
ThursdayN =""
FridayM =""
SaturdayN =""
# Pavel_st = 0
# Pavel_fr = 0
NewSat = ""

i = 0
count = 0
s = 0


# def button_callback():
#     global val
#     val = (" " + val + (optionmenu_1.get()) + " " + (combobox_1.get()) + " ")
#     label_2 = customtkinter.CTkLabel(master=frame_1,
#                                      text=" " + (optionmenu_1.get()) + " can not work on " + (combobox_1.get()) + " ")
#     label_2.grid(pady=1, padx=1)

def check():
    global SundayM
    global SundayN
    global MondayM
    global MondayN
    global TuesdayM
    global TuesdayN
    global WednesdayM
    global WednesdayN
    global ThursdayM
    global ThursdayN
    global FridayM
    global SaturdayN
    block()
    if not "Alyona All" in val:
        if not "Alyona Sun-Morning" in val:
            SundayM = SundayM + "Alyona "
        if not  "Alyona Mon-Morning" in val:
            MondayM = MondayM + "Alyona "
        if not "Alyona Tue-Morning" in val:
            TuesdayM = TuesdayM + "Alyona "
        if not "Alyona Wed-Morning" in val:
            WednesdayM = WednesdayM + "Alyona "
        if not "Alyona Thu-Morning" in val:
            ThursdayM = ThursdayM + "Alyona "
        if not "Alyona Fri-Morning " in val:
            FridayM = FridayM + "Alyona "
        if not "Alyona Sun-Night" in val:
            SundayN = SundayN + "Alyona "
        if not "Alyona Mon-Night" in val:
            MondayN = MondayN + "Alyona "
        if not "Alyona Tue-Night" in val:
            TuesdayN = TuesdayN + "Alyona "
        if not "Alyona Wed-Night" in val:
            WednesdayN = WednesdayN + "Alyona "
        if not "Alyona Thu-Night" in val:
            ThursdayN = ThursdayN + "Alyona "
        if not "Alyona Sat-Night " in val:
            SaturdayN = SaturdayN + "Alyona "
    if not "Alex All" in val:
        if not "Alex Sun-Morning" in val:
            SundayM = SundayM + "Alex "
        if not "Alex Mon-Morning" in val:
            MondayM = MondayM + "Alex "
        if not "Alex Tue-Morning" in val:
            TuesdayM = TuesdayM + "Alex "
        if not "Alex Wed-Morning" in val:
            WednesdayM = WednesdayM + "Alex "
        if not "Alex Thu-Morning" in val:
            ThursdayM = ThursdayM + "Alex "
        if not "Alex Fri-Morning " in val:
            FridayM = FridayM + "Alex "
        if not "Alex Sun-Night" in val:
            SundayN = SundayN + "Alex "
        if not "Alex Mon-Night" in val:
            MondayN = MondayN + "Alex "
        if not "Alex Tue-Night" in val:
            TuesdayN = TuesdayN + "Alex "
        if not "Alex Wed-Night" in val:
            WednesdayN = WednesdayN + "Alex "
        if not "Alex Thu-Night" in val:
            ThursdayN = ThursdayN + "Alex "
        if not "Alex Sat-Night " in val:
            SaturdayN = SaturdayN + "Alex "
    if not "Ofir All" in val:
        if not "Ofir Sun-Morning" in val:
            SundayM = SundayM + "Ofir "
        if not "Ofir Mon-Morning" in val:
            MondayM = MondayM + "Ofir "
        if not "Ofir Tue-Morning" in val:
            TuesdayM = TuesdayM + "Ofir "
        if not "Ofir Wed-Morning" in val:
            WednesdayM = WednesdayM + "Ofir "
        if not "Ofir Thu-Morning" in val:
            ThursdayM = ThursdayM + "Ofir "
        if not "Ofir Fri-Morning " in val:
            FridayM = FridayM + "Ofir "
        if not "Ofir Sun-Night" in val:
            SundayN = SundayN + "Ofir "
        if not "Ofir Mon-Night" in val:
            MondayN = MondayN + "Ofir "
        if not "Ofir Tue-Night" in val:
            TuesdayN = TuesdayN + "Ofir "
        if not "Ofir Wed-Night" in val:
            WednesdayN = WednesdayN + "Ofir "
        if not "Ofir Thu-Night" in val:
            ThursdayN = ThursdayN + "Ofir "
        if not "Ofir Sat-Night " in val:
            SaturdayN = SaturdayN + "Ofir "
    if not "Almog All" in val:
        if not "Almog Sun-Morning" in val:
            SundayM = SundayM + "Almog "
        if not "Almog Mon-Morning" in val:
            MondayM = MondayM + "Almog "
        if not "Almog Tue-Morning" in val:
            TuesdayM = TuesdayM + "Almog "
        if not "Almog Wed-Morning" in val:
            WednesdayM = WednesdayM + "Almog "
        if not "Almog Thu-Morning" in val:
            ThursdayM = ThursdayM + "Almog "
        if not "Almog Fri-Morning " in val:
            FridayM = FridayM + "Almog "
        if not "Almog Sun-Night" in val:
            SundayN = SundayN + "Almog "
        if not "Almog Mon-Night" in val:
            MondayN = MondayN + "Almog "
        if not "Almog Tue-Night" in val:
            TuesdayN = TuesdayN + "Almog "
        if not "Almog Wed-Night" in val:
            WednesdayN = WednesdayN + "Almog "
        if not "Almog Thu-Night" in val:
            ThursdayN = ThursdayN + "Almog "
        if not "Almog Sat-Night " in val:
            SaturdayN = SaturdayN + "Almog "
    if not "Ran All" in val:
        if not "Ran Sun-Morning" in val:
            SundayM = SundayM + "Ran "
        if not "Ran Mon-Morning" in val:
            MondayM = MondayM + "Ran "
        if not "Ran Tue-Morning" in val:
            TuesdayM = TuesdayM + "Ran "
        if not "Ran Wed-Morning" in val:
            WednesdayM = WednesdayM + "Ran "
        if not "Ran Thu-Morning" in val:
            ThursdayM = ThursdayM + "Ran "
        if not "Ran Fri-Morning " in val:
            FridayM = FridayM + "Ran "
        if not "Ran Sun-Night" in val:
            SundayN = SundayN + "Ran "
        if not "Ran Mon-Night" in val:
            MondayN = MondayN + "Ran "
        if not "Ran Tue-Night" in val:
            TuesdayN = TuesdayN + "Ran "
        if not "Ran Wed-Night" in val:
            WednesdayN = WednesdayN + "Ran "
        if not "Ran Thu-Night" in val:
            ThursdayN = ThursdayN + "Ran "
        if not "Ran Sat-Night " in val:
            SaturdayN = SaturdayN + "Ran "
    if not "Sahar All" in val:
        if not "Sahar Sun-Morning" in val:
            SundayM = SundayM + "Sahar "
        if not "Sahar Mon-Morning" in val:
            MondayM = MondayM + "Sahar "
        if not "Sahar Tue-Morning" in val:
            TuesdayM = TuesdayM + "Sahar "
        if not "Sahar Wed-Morning" in val:
            WednesdayM = WednesdayM + "Sahar "
        if not "Sahar Thu-Morning" in val:
            ThursdayM = ThursdayM + "Sahar "
        if not "Sahar Fri-Morning " in val:
            FridayM = FridayM + "Sahar "
        if not "Sahar Sun-Night" in val:
            SundayN = SundayN + "Sahar "
        if not "Sahar Mon-Night" in val:
            MondayN = MondayN + "Sahar "
        if not "Sahar Tue-Night" in val:
            TuesdayN = TuesdayN + "Sahar "
        if not "Sahar Wed-Night" in val:
            WednesdayN = WednesdayN + "Sahar "
        if not "Sahar Thu-Night" in val:
            ThursdayN = ThursdayN + "Sahar "
        if not "Sahar Sat-Night " in val:
            SaturdayN = SaturdayN + "Sahar "
    if not "Yair All" in val:
        if not "Yair Sun-Morning" in val:
            SundayM = SundayM + "Yair "
        if not "Yair Mon-Morning" in val:
            MondayM = MondayM + "Yair "
        if not "Yair Tue-Morning" in val:
            TuesdayM = TuesdayM + "Yair "
        if not "Yair Wed-Morning" in val:
            WednesdayM = WednesdayM + "Yair "
        if not "Yair Thu-Morning" in val:
            ThursdayM = ThursdayM + "Yair "
        if not "Yair Fri-Morning " in val:
            FridayM = FridayM + "Yair "
        if not "Yair Sun-Night" in val:
            SundayN = SundayN + "Yair "
        if not "Yair Mon-Night" in val:
            MondayN = MondayN + "Yair "
        if not "Yair Tue-Night" in val:
            TuesdayN = TuesdayN + "Yair "
        if not "Yair Wed-Night" in val:
            WednesdayN = WednesdayN + "Yair "
        if not "Yair Thu-Night" in val:
            ThursdayN = ThursdayN + "Yair "
        if not "Yair Sat-Night " in val:
            SaturdayN = SaturdayN + "Yair "
    #trasform into List
    SundayM = SundayM.split(" ")
    SundayM.pop()

    SundayN = SundayN.split(" ")
    SundayN.pop()

    MondayM = MondayM.split(" ")
    MondayM.pop()

    MondayN = MondayN.split(" ")
    MondayN.pop()

    TuesdayM = TuesdayM.split(" ")
    TuesdayM.pop()

    TuesdayN = TuesdayN.split(" ")
    TuesdayN.pop()

    ThursdayM = ThursdayM.split(" ")
    ThursdayM.pop()

    ThursdayN = ThursdayN.split(" ")
    ThursdayN.pop()

    WednesdayM = WednesdayM.split(" ")
    WednesdayM.pop()

    WednesdayN = WednesdayN.split(" ")
    WednesdayN.pop()

    FridayM = FridayM.split(" ")
    FridayM.pop()

    SaturdayN = SaturdayN.split(" ")
    SaturdayN.pop()



    print("hello")



def button_callback2():
    global SundayM
    global SundayN
    global MondayM
    global MondayN
    global TuesdayM
    global TuesdayN
    global WednesdayM
    global WednesdayN
    global ThursdayM
    global ThursdayN
    global FridayM
    global SaturdayN
    global val
    global hello
    global Alyona_st
    global Alex_st
    global Ofir_st
    global Yair_st
    global Almog_st
    global Ran_st
    global Sahar_st
    global Alyona_fr
    global Alex_fr
    global Ofir_fr
    global Yair_fr
    global Almog_fr
    global Ran_fr
    global Sahar_fr

    fr = [Alyona_fr, Alex_fr, Ofir_fr, Almog_fr,  Yair_fr, Sahar_fr, Ran_fr]
    st = [Alyona_st, Alex_st, Ofir_st, Almog_st,  Yair_st, Sahar_st, Ran_st]

    # after get from input which shifts each ops can not work, tRansform the data to format 'X = (can not work) ABC..'
    check()
    print("SundayM:", SundayM ,"\n",
          "SundayN:", SundayN ,"\n",
          "MondayM:", MondayM ,"\n",
          "MondayN:", MondayN ,"\n",
          "TuesdayM:", TuesdayM ,"\n",
          "TuesdayN:", TuesdayN ,"\n",
          "WednesdayM:", WednesdayM ,"\n",
          "WednesdayN:", WednesdayN ,"\n",
          "ThursdayM:", ThursdayM ,"\n",
          "ThursdayN:", ThursdayN ,"\n",
          "FridayM:", FridayM ,"\n",
          "SaturdayN:", SaturdayN) ,"\n"

    #sunday
    global newSat
    for i in range(0,3):
        if len(SaturdayN) >= 2:
            if (Alyona_st <= min(st) + i) and ("Alyona" in SaturdayN):
                NewSat = "Alyona"
            elif (Ofir_st <= min(st) + i) and ("Ofir" in SaturdayN):
                NewSat = "Ofir"
            elif (Yair_st <= min(st) + i) and ("Yair" in SaturdayN):
                NewSat = "Yair"
            elif (Sahar_st <= min(st) + i) and ("Sahar" in SaturdayN):
                NewSat = "Sahar"
            elif (Alex_st <= min(st) + i) and ("Alex" in SaturdayN):
                NewSat = "Alex"
            elif (Ran_st <= min(st) + i) and ("Ran" in SaturdayN):
                NewSat = "Ran"
            elif (Almog_st <= min(st) + i) and ("Almog" in SaturdayN):
                NewSat = "Almog"
        if len(SaturdayN) >= 2:
            if (Alyona_st <= min(st) + i) and ("Alyona" in SaturdayN) and (not "Alyona" in NewSat):
                NewSat[1] = "Alyona"
            elif (Ofir_st <= min(st) + i) and ("Ofir" in SaturdayN) and (not "Ofir" in NewSat):
                NewSat[1] = "Ofir"
            elif (Yair_st <= min(st) + i) and ("Yair" in SaturdayN) and (not "Yair" in NewSat):
                NewSat[1] = "Yair"
            elif (Sahar_st <= min(st) + i) and ("Sahar" in SaturdayN) and (not "Sahar" in NewSat):
                NewSat[1] = "Sahar"
            elif (Alex_st <= min(st) + i) and ("Alex" in SaturdayN) and (not "Alex" in NewSat):
                NewSat[1] = "Alex"
            elif (Ran_st <= min(st) + i) and ("Ran" in SaturdayN) and (not "Ran" in NewSat):
                NewSat[1] = "Ran"
            elif (Almog_st <= min(st) + i) and ("Almog" in SaturdayN) and (not "Almog" in NewSat):
                NewSat[1] = "Almog"



    if len(SundayN) >= 3:
        SundayN = random.sample(SundayN, 3)
        if SundayN[0] in SundayM: SundayM.remove(SundayN[0])
        if SundayN[1] in SundayM: SundayM.remove(SundayN[1])
        if SundayN[2] in SundayM: SundayM.remove(SundayN[2])

        if SundayN[0] in MondayM: MondayM.remove(SundayN[0])
        if SundayN[1] in MondayM: MondayM.remove(SundayN[1])
        if SundayN[2] in MondayM: MondayM.remove(SundayN[2])
    elif len(SundayN) == 2:
        if SundayN[0] in SundayM: SundayM.remove(SundayN[0])
        if SundayN[1] in SundayM: SundayM.remove(SundayN[1])

        if SundayN[0] in MondayM: MondayM.remove(SundayN[0])
        if SundayN[1] in MondayM: MondayM.remove(SundayN[1])
    elif len(SundayN) == 1:
        if SundayN[0] in SundayM: SundayM.remove(SundayN[0])
        if SundayN[0] in MondayM: MondayM.remove(SundayN[0])

    if len(MondayN) >= 3:
        MondayN = random.sample(MondayN, 3)
        if MondayN[0] in TuesdayM: TuesdayM.remove(MondayN[0])
        if MondayN[1] in TuesdayM: TuesdayM.remove(MondayN[1])
        if MondayN[2] in TuesdayM: TuesdayM.remove(MondayN[2])

        if MondayN[0] in MondayM: MondayM.remove(MondayN[0])
        if MondayN[1] in MondayM: MondayM.remove(MondayN[1])
        if MondayN[2] in MondayM: MondayM.remove(MondayN[2])
    elif len(MondayN) == 2:
        if MondayN[0] in TuesdayM: TuesdayM.remove(MondayN[0])
        if MondayN[1] in TuesdayM: TuesdayM.remove(MondayN[1])

        if MondayN[0] in MondayM: MondayM.remove(MondayN[0])
        if MondayN[1] in MondayM: MondayM.remove(MondayN[1])
    elif len(MondayN) == 1:
        if MondayN[0] in TuesdayM: TuesdayM.remove(MondayN[0])
        if MondayN[0] in MondayM: MondayM.remove(MondayN[0])

    if len(TuesdayN) >= 3:
        TuesdayN = random.sample(TuesdayN, 3)
        if TuesdayN[0] in WednesdayM: WednesdayM.remove(TuesdayN[0])
        if TuesdayN[1] in WednesdayM: WednesdayM.remove(TuesdayN[1])
        if TuesdayN[2] in WednesdayM: WednesdayM.remove(TuesdayN[2])

        if TuesdayN[0] in TuesdayM: TuesdayM.remove(TuesdayN[0])
        if TuesdayN[1] in TuesdayM: TuesdayM.remove(TuesdayN[1])
        if TuesdayN[2] in TuesdayM: TuesdayM.remove(TuesdayN[2])
    elif len(TuesdayN) == 2:
        if TuesdayN[0] in WednesdayM: WednesdayM.remove(TuesdayN[0])
        if TuesdayN[1] in WednesdayM: WednesdayM.remove(TuesdayN[1])

        if TuesdayN[0] in TuesdayM: TuesdayM.remove(TuesdayN[0])
        if TuesdayN[1] in TuesdayM: TuesdayM.remove(TuesdayN[1])
    elif len(TuesdayN) == 1:
        if TuesdayN[0] in WednesdayM: WednesdayM.remove(TuesdayN[0])
        if TuesdayN[0] in TuesdayM: TuesdayM.remove(TuesdayN[0])


    print("New SundayM:", SundayM ,"\n",
          "New SundayN:", SundayN ,"\n",
          "New MondayM:", MondayM ,"\n",
          "New MondayN:", MondayN ,"\n",
          "New TuesdayM:", TuesdayM ,"\n",
          "New TuesdayN:", TuesdayN ,"\n",
          "New WednesdayM:", WednesdayM ,"\n",
          "New WednesdayN:", WednesdayN ,"\n",
          "New ThursdayM:", ThursdayM ,"\n",
          "New ThursdayN:", ThursdayN ,"\n",
          "New FridayM:", FridayM ,"\n",
          "New SaturdayN:", SaturdayN) ,"\n"






    # his()
    # design()
    # absolutePath = Path('../SC/shifts.xlsx').resolve()
    # os.system(f'start excel.exe "{absolutePath}"')

def design():
    global val
    global hello
    # excel write
    ws['A40'].style = highlight
    ws['A2'] = (hello[0])
    ws['A3'] = (hello[1])
    ws['A4'] = (hello[2])
    ws['A5'] = (hello[3])
    ws['C2'] = (hello[0])
    ws['C3'] = (hello[1])
    ws['C4'] = (hello[2])
    ws['C5'] = (hello[3])
    ws['E2'] = (hello[0])
    ws['E3'] = (hello[1])
    ws['F4'] = (hello[2])
    ws['F5'] = (hello[3])
    ws['H2'] = (hello[0])
    ws['G3'] = (hello[1])
    ws['H4'] = (hello[2])
    ws['H5'] = (hello[3])
    ws['J2'] = (hello[0])
    ws['J4'] = (hello[2])
    ws['J5'] = (hello[3])
    ws['B6'] = (hello[4])
    ws['B7'] = (hello[5])
    ws['C8'] = (hello[6])
    ws['B9'] = (hello[7])
    ws['B10'] = (hello[8])
    ws['B11'] = (hello[9])
    ws['D6'] = (hello[4])
    ws['D7'] = (hello[5])
    ws['E8'] = (hello[6])
    ws['D9'] = (hello[7])
    ws['E10'] = (hello[8])
    ws['E11'] = (hello[9])
    ws['G6'] = (hello[4])
    ws['F7'] = (hello[5])
    ws['H8'] = (hello[6])
    ws['F9'] = (hello[7])
    ws['G10'] = (hello[8])
    ws['G11'] = (hello[9])
    ws['I3'] = (hello[1])
    ws['I6'] = (hello[4])
    ws['I7'] = (hello[5])
    ws['J8'] = (hello[6])
    ws['I9'] = (hello[7])
    ws['J10'] = (hello[8])
    ws['I11'] = (hello[9])
    ws['L6'] = (hello[4])
    ws['K7'] = (hello[5])
    ws['L8'] = (hello[6])
    ws['K9'] = (hello[7])
    ws['L10'] = (hello[8])
    ws['K11'] = (hello[9])

    if 'AlmogVacSun' in val:
        if 'Almog' in hello[0]:
            ws['A2'] = "Vac"
            ws['B2'] = "Vac"
        if 'Almog' in hello[1]:
            ws['A3'] = "Vac"
            ws['B3'] = "Vac"
        if 'Almog' in hello[2]:
            ws['A4'] = "Vac"
            ws['B4'] = "Vac"
        if 'Almog' in hello[3]:
            ws['A5'] = "Vac"
            ws['B5'] = "Vac"
        if 'Almog' in hello[4]:
            ws['A6'] = "Vac"
            ws['B6'] = "Vac"
        if 'Almog' in hello[5]:
            ws['A7'] = "Vac"
            ws['B7'] = "Vac"
        if 'Almog' in hello[6]:
            ws['A8'] = "Vac"
            ws['B8'] = "Vac"
        if 'Almog' in hello[7]:
            ws['A9'] = "Vac"
            ws['B9'] = "Vac"
        if 'Almog' in hello[8]:
            ws['A10'] = "Vac"
            ws['B10'] = "Vac"
        if 'Almog' in hello[9]:
            ws['A11'] = "Vac"
            ws['B11'] = "Vac"
    if 'AlmogVacMon' in val:
        if 'Almog' in hello[0]:
            ws['C2'] = "Vac"
            ws['D2'] = "Vac"
        if 'Almog' in hello[1]:
            ws['C3'] = "Vac"
            ws['D3'] = "Vac"
        if 'Almog' in hello[2]:
            ws['C4'] = "Vac"
            ws['D4'] = "Vac"
        if 'Almog' in hello[3]:
            ws['C5'] = "Vac"
            ws['D5'] = "Vac"
        if 'Almog' in hello[4]:
            ws['C6'] = "Vac"
            ws['D6'] = "Vac"
        if 'Almog' in hello[5]:
            ws['C7'] = "Vac"
            ws['D7'] = "Vac"
        if 'Almog' in hello[6]:
            ws['C8'] = "Vac"
            ws['D8'] = "Vac"
        if 'Almog' in hello[7]:
            ws['C9'] = "Vac"
            ws['D9'] = "Vac"
        if 'Almog' in hello[8]:
            ws['C10'] = "Vac"
            ws['D10'] = "Vac"
        if 'Almog' in hello[9]:
            ws['C11'] = "Vac"
            ws['D11'] = "Vac"
    if 'AlmogVacTue' in val:
        if 'Almog' in hello[0]:
            ws['E2'] = "Vac"
            ws['F2'] = "Vac"
        if 'Almog' in hello[1]:
            ws['E3'] = "Vac"
            ws['F3'] = "Vac"
        if 'Almog' in hello[2]:
            ws['E4'] = "Vac"
            ws['F4'] = "Vac"
        if 'Almog' in hello[3]:
            ws['E5'] = "Vac"
            ws['F5'] = "Vac"
        if 'Almog' in hello[4]:
            ws['E6'] = "Vac"
            ws['F6'] = "Vac"
        if 'Almog' in hello[5]:
            ws['E7'] = "Vac"
            ws['F7'] = "Vac"
        if 'Almog' in hello[6]:
            ws['E8'] = "Vac"
            ws['F8'] = "Vac"
        if 'Almog' in hello[7]:
            ws['E9'] = "Vac"
            ws['F9'] = "Vac"
        if 'Almog' in hello[8]:
            ws['E10'] = "Vac"
            ws['F10'] = "Vac"
        if 'Almog' in hello[9]:
            ws['E11'] = "Vac"
            ws['F11'] = "Vac"
    if 'AlmogVacWed' in val:
        if 'Almog' in hello[0]:
            ws['G2'] = "Vac"
            ws['H2'] = "Vac"
        if 'Almog' in hello[1]:
            ws['G3'] = "Vac"
            ws['H3'] = "Vac"
        if 'Almog' in hello[2]:
            ws['G4'] = "Vac"
            ws['H4'] = "Vac"
        if 'Almog' in hello[3]:
            ws['G5'] = "Vac"
            ws['H5'] = "Vac"
        if 'Almog' in hello[4]:
            ws['G6'] = "Vac"
            ws['H6'] = "Vac"
        if 'Almog' in hello[5]:
            ws['G7'] = "Vac"
            ws['H7'] = "Vac"
        if 'Almog' in hello[6]:
            ws['G8'] = "Vac"
            ws['H8'] = "Vac"
        if 'Almog' in hello[7]:
            ws['G9'] = "Vac"
            ws['H9'] = "Vac"
        if 'Almog' in hello[8]:
            ws['G10'] = "Vac"
            ws['H10'] = "Vac"
        if 'Almog' in hello[9]:
            ws['G11'] = "Vac"
            ws['H11'] = "Vac"
    if 'AlmogVacThu' in val:
        if 'Almog' in hello[0]:
            ws['I2'] = "Vac"
            ws['j2'] = "Vac"
        if 'Almog' in hello[1]:
            ws['I3'] = "Vac"
            ws['j3'] = "Vac"
        if 'Almog' in hello[2]:
            ws['I4'] = "Vac"
            ws['j4'] = "Vac"
        if 'Almog' in hello[3]:
            ws['I5'] = "Vac"
            ws['j5'] = "Vac"
        if 'Almog' in hello[4]:
            ws['I6'] = "Vac"
            ws['j6'] = "Vac"
        if 'Almog' in hello[5]:
            ws['I7'] = "Vac"
            ws['j7'] = "Vac"
        if 'Almog' in hello[6]:
            ws['I8'] = "Vac"
            ws['j8'] = "Vac"
        if 'Almog' in hello[7]:
            ws['I9'] = "Vac"
            ws['j9'] = "Vac"
        if 'Almog' in hello[8]:
            ws['I10'] = "Vac"
            ws['j10'] = "Vac"
        if 'Almog' in hello[9]:
            ws['I11'] = "Vac"
            ws['j11'] = "Vac"
    if 'AlmogVacFri' in val:
        if 'Almog' in hello[0]:
            ws['K2'] = "Vac"
        if 'Almog' in hello[1]:
            ws['K3'] = "Vac"
        if 'Almog' in hello[2]:
            ws['K4'] = "Vac"
        if 'Almog' in hello[3]:
            ws['K5'] = "Vac"
        if 'Almog' in hello[4]:
            ws['K6'] = "Vac"
        if 'Almog' in hello[5]:
            ws['K7'] = "Vac"
        if 'Almog' in hello[6]:
            ws['K8'] = "Vac"
        if 'Almog' in hello[7]:
            ws['K9'] = "Vac"
        if 'Almog' in hello[8]:
            ws['K10'] = "Vac"
        if 'Almog' in hello[9]:
            ws['K11'] = "Vac"
    if 'AlmogVacSat' in val:
        if 'Almog' in hello[0]:
            ws['L2'] = "Vac"
        if 'Almog' in hello[1]:
            ws['L3'] = "Vac"
        if 'Almog' in hello[2]:
            ws['L4'] = "Vac"
        if 'Almog' in hello[3]:
            ws['L5'] = "Vac"
        if 'Almog' in hello[4]:
            ws['L6'] = "Vac"
        if 'Almog' in hello[5]:
            ws['L7'] = "Vac"
        if 'Almog' in hello[6]:
            ws['L8'] = "Vac"
        if 'Almog' in hello[7]:
            ws['L9'] = "Vac"
        if 'Almog' in hello[8]:
            ws['L10'] = "Vac"
        if 'Almog' in hello[9]:
            ws['L11'] = "Vac"
    if 'AlmogVacAll' in val:
        if 'Almog' in hello[0]:
            ws['A2'] = "Almog-Vac"
            ws['B2'] = "Almog-Vac"
            ws['C2'] = "Almog-Vac"
            ws['D2'] = "Almog-Vac"
            ws['E2'] = "Almog-Vac"
            ws['F2'] = "Almog-Vac"
            ws['G2'] = "Almog-Vac"
            ws['H2'] = "Almog-Vac"
            ws['I2'] = "Almog-Vac"
            ws['J2'] = "Almog-Vac"
            ws['K2'] = "Almog-Vac"
            ws['L2'] = "Almog-Vac"
        if 'Almog' in hello[1]:
            ws['A3'] = "Almog-Vac"
            ws['B3'] = "Almog-Vac"
            ws['C3'] = "Almog-Vac"
            ws['D3'] = "Almog-Vac"
            ws['E3'] = "Almog-Vac"
            ws['F3'] = "Almog-Vac"
            ws['G3'] = "Almog-Vac"
            ws['H3'] = "Almog-Vac"
            ws['I3'] = "Almog-Vac"
            ws['J3'] = "Almog-Vac"
            ws['K3'] = "Almog-Vac"
            ws['L3'] = "Almog-Vac"
        if 'Almog' in hello[2]:
            ws['A4'] = "Almog-Vac"
            ws['B4'] = "Almog-Vac"
            ws['C4'] = "Almog-Vac"
            ws['D4'] = "Almog-Vac"
            ws['E4'] = "Almog-Vac"
            ws['F4'] = "Almog-Vac"
            ws['G4'] = "Almog-Vac"
            ws['H4'] = "Almog-Vac"
            ws['I4'] = "Almog-Vac"
            ws['J4'] = "Almog-Vac"
            ws['K4'] = "Almog-Vac"
            ws['L4'] = "Almog-Vac"
        if 'Almog' in hello[3]:
            ws['A5'] = "Almog-Vac"
            ws['B5'] = "Almog-Vac"
            ws['C5'] = "Almog-Vac"
            ws['D5'] = "Almog-Vac"
            ws['E5'] = "Almog-Vac"
            ws['F5'] = "Almog-Vac"
            ws['G5'] = "Almog-Vac"
            ws['H5'] = "Almog-Vac"
            ws['I5'] = "Almog-Vac"
            ws['J5'] = "Almog-Vac"
            ws['K5'] = "Almog-Vac"
            ws['L5'] = "Almog-Vac"
        if 'Almog' in hello[4]:
            ws['A6'] = "Almog-Vac"
            ws['B6'] = "Almog-Vac"
            ws['C6'] = "Almog-Vac"
            ws['D6'] = "Almog-Vac"
            ws['E6'] = "Almog-Vac"
            ws['F6'] = "Almog-Vac"
            ws['G6'] = "Almog-Vac"
            ws['H6'] = "Almog-Vac"
            ws['I6'] = "Almog-Vac"
            ws['J6'] = "Almog-Vac"
            ws['K6'] = "Almog-Vac"
            ws['L6'] = "Almog-Vac"
        if 'Almog' in hello[5]:
            ws['A7'] = "Almog-Vac"
            ws['B7'] = "Almog-Vac"
            ws['C7'] = "Almog-Vac"
            ws['D7'] = "Almog-Vac"
            ws['E7'] = "Almog-Vac"
            ws['F7'] = "Almog-Vac"
            ws['G7'] = "Almog-Vac"
            ws['H7'] = "Almog-Vac"
            ws['I7'] = "Almog-Vac"
            ws['J7'] = "Almog-Vac"
            ws['K7'] = "Almog-Vac"
            ws['L7'] = "Almog-Vac"
        if 'Almog' in hello[6]:
            ws['A8'] = "Almog-Vac"
            ws['B8'] = "Almog-Vac"
            ws['C8'] = "Almog-Vac"
            ws['D8'] = "Almog-Vac"
            ws['E8'] = "Almog-Vac"
            ws['F8'] = "Almog-Vac"
            ws['G8'] = "Almog-Vac"
            ws['H8'] = "Almog-Vac"
            ws['I8'] = "Almog-Vac"
            ws['J8'] = "Almog-Vac"
            ws['K8'] = "Almog-Vac"
            ws['L8'] = "Almog-Vac"
        if 'Almog' in hello[7]:
            ws['A9'] = "Almog-Vac"
            ws['B9'] = "Almog-Vac"
            ws['C9'] = "Almog-Vac"
            ws['D9'] = "Almog-Vac"
            ws['E9'] = "Almog-Vac"
            ws['F9'] = "Almog-Vac"
            ws['G9'] = "Almog-Vac"
            ws['H9'] = "Almog-Vac"
            ws['I9'] = "Almog-Vac"
            ws['J9'] = "Almog-Vac"
            ws['K9'] = "Almog-Vac"
            ws['L9'] = "Almog-Vac"
        if 'Almog' in hello[8]:
            ws['A9'] = "Almog-Vac"
            ws['B9'] = "Almog-Vac"
            ws['C9'] = "Almog-Vac"
            ws['D9'] = "Almog-Vac"
            ws['E9'] = "Almog-Vac"
            ws['F9'] = "Almog-Vac"
            ws['G9'] = "Almog-Vac"
            ws['H9'] = "Almog-Vac"
            ws['I9'] = "Almog-Vac"
            ws['J9'] = "Almog-Vac"
            ws['K9'] = "Almog-Vac"
            ws['L9'] = "Almog-Vac"
        if 'Almog' in hello[9]:
            ws['A10'] = "Almog-Vac"
            ws['B10'] = "Almog-Vac"
            ws['C10'] = "Almog-Vac"
            ws['D10'] = "Almog-Vac"
            ws['E10'] = "Almog-Vac"
            ws['F10'] = "Almog-Vac"
            ws['G10'] = "Almog-Vac"
            ws['H10'] = "Almog-Vac"
            ws['I10'] = "Almog-Vac"
            ws['J10'] = "Almog-Vac"
            ws['K10'] = "Almog-Vac"
            ws['L10'] = "Almog-Vac"
    if 'AlyonaVacSun' in val:
        if 'Alyona' in hello[0]:
            ws['A2'] = "Vac"
            ws['B2'] = "Vac"
        if 'Alyona' in hello[1]:
            ws['A3'] = "Vac"
            ws['B3'] = "Vac"
        if 'Alyona' in hello[2]:
            ws['A4'] = "Vac"
            ws['B4'] = "Vac"
        if 'Alyona' in hello[3]:
            ws['A5'] = "Vac"
            ws['B5'] = "Vac"
        if 'Alyona' in hello[4]:
            ws['A6'] = "Vac"
            ws['B6'] = "Vac"
        if 'Alyona' in hello[5]:
            ws['A7'] = "Vac"
            ws['B7'] = "Vac"
        if 'Alyona' in hello[6]:
            ws['A8'] = "Vac"
            ws['B8'] = "Vac"
        if 'Alyona' in hello[7]:
            ws['A9'] = "Vac"
            ws['B9'] = "Vac"
        if 'Alyona' in hello[8]:
            ws['A10'] = "Vac"
            ws['B10'] = "Vac"
        if 'Alyona' in hello[9]:
            ws['A11'] = "Vac"
            ws['B11'] = "Vac"
    if 'AlyonaVacMon' in val:
        if 'Alyona' in hello[0]:
            ws['C2'] = "Vac"
            ws['D2'] = "Vac"
        if 'Alyona' in hello[1]:
            ws['C3'] = "Vac"
            ws['D3'] = "Vac"
        if 'Alyona' in hello[2]:
            ws['C4'] = "Vac"
            ws['D4'] = "Vac"
        if 'Alyona' in hello[3]:
            ws['C5'] = "Vac"
            ws['D5'] = "Vac"
        if 'Alyona' in hello[4]:
            ws['C6'] = "Vac"
            ws['D6'] = "Vac"
        if 'Alyona' in hello[5]:
            ws['C7'] = "Vac"
            ws['D7'] = "Vac"
        if 'Alyona' in hello[6]:
            ws['C8'] = "Vac"
            ws['D8'] = "Vac"
        if 'Alyona' in hello[7]:
            ws['C9'] = "Vac"
            ws['D9'] = "Vac"
        if 'Alyona' in hello[8]:
            ws['C10'] = "Vac"
            ws['D10'] = "Vac"
        if 'Alyona' in hello[9]:
            ws['C11'] = "Vac"
            ws['D11'] = "Vac"
    if 'AlyonaVacTue' in val:
        if 'Alyona' in hello[0]:
            ws['E2'] = "Vac"
            ws['F2'] = "Vac"
        if 'Alyona' in hello[1]:
            ws['E3'] = "Vac"
            ws['F3'] = "Vac"
        if 'Alyona' in hello[2]:
            ws['E4'] = "Vac"
            ws['F4'] = "Vac"
        if 'Alyona' in hello[3]:
            ws['E5'] = "Vac"
            ws['F5'] = "Vac"
        if 'Alyona' in hello[4]:
            ws['E6'] = "Vac"
            ws['F6'] = "Vac"
        if 'Alyona' in hello[5]:
            ws['E7'] = "Vac"
            ws['F7'] = "Vac"
        if 'Alyona' in hello[6]:
            ws['E8'] = "Vac"
            ws['F8'] = "Vac"
        if 'Alyona' in hello[7]:
            ws['E9'] = "Vac"
            ws['F9'] = "Vac"
        if 'Alyona' in hello[8]:
            ws['E10'] = "Vac"
            ws['F10'] = "Vac"
        if 'Alyona' in hello[9]:
            ws['E11'] = "Vac"
            ws['F11'] = "Vac"
    if 'AlyonaVacWed' in val:
        if 'Alyona' in hello[0]:
            ws['G2'] = "Vac"
            ws['H2'] = "Vac"
        if 'Alyona' in hello[1]:
            ws['G3'] = "Vac"
            ws['H3'] = "Vac"
        if 'Alyona' in hello[2]:
            ws['G4'] = "Vac"
            ws['H4'] = "Vac"
        if 'Alyona' in hello[3]:
            ws['G5'] = "Vac"
            ws['H5'] = "Vac"
        if 'Alyona' in hello[4]:
            ws['G6'] = "Vac"
            ws['H6'] = "Vac"
        if 'Alyona' in hello[5]:
            ws['G7'] = "Vac"
            ws['H7'] = "Vac"
        if 'Alyona' in hello[6]:
            ws['G8'] = "Vac"
            ws['H8'] = "Vac"
        if 'Alyona' in hello[7]:
            ws['G9'] = "Vac"
            ws['H9'] = "Vac"
        if 'Alyona' in hello[8]:
            ws['G10'] = "Vac"
            ws['H10'] = "Vac"
        if 'Alyona' in hello[9]:
            ws['G11'] = "Vac"
            ws['H11'] = "Vac"
    if 'AlyonaVacThu' in val:
        if 'Alyona' in hello[0]:
            ws['I2'] = "Vac"
            ws['j2'] = "Vac"
        if 'Alyona' in hello[1]:
            ws['I3'] = "Vac"
            ws['j3'] = "Vac"
        if 'Alyona' in hello[2]:
            ws['I4'] = "Vac"
            ws['j4'] = "Vac"
        if 'Alyona' in hello[3]:
            ws['I5'] = "Vac"
            ws['j5'] = "Vac"
        if 'Alyona' in hello[4]:
            ws['I6'] = "Vac"
            ws['j6'] = "Vac"
        if 'Alyona' in hello[5]:
            ws['I7'] = "Vac"
            ws['j7'] = "Vac"
        if 'Alyona' in hello[6]:
            ws['I8'] = "Vac"
            ws['j8'] = "Vac"
        if 'Alyona' in hello[7]:
            ws['I9'] = "Vac"
            ws['j9'] = "Vac"
        if 'Alyona' in hello[8]:
            ws['I10'] = "Vac"
            ws['j10'] = "Vac"
        if 'Alyona' in hello[9]:
            ws['I11'] = "Vac"
            ws['j11'] = "Vac"
    if 'AlyonaVacFri' in val:
        if 'Alyona' in hello[0]:
            ws['K2'] = "Vac"
        if 'Alyona' in hello[1]:
            ws['K3'] = "Vac"
        if 'Alyona' in hello[2]:
            ws['K4'] = "Vac"
        if 'Alyona' in hello[3]:
            ws['K5'] = "Vac"
        if 'Alyona' in hello[4]:
            ws['K6'] = "Vac"
        if 'Alyona' in hello[5]:
            ws['K7'] = "Vac"
        if 'Alyona' in hello[6]:
            ws['K8'] = "Vac"
        if 'Alyona' in hello[7]:
            ws['K9'] = "Vac"
        if 'Alyona' in hello[8]:
            ws['K10'] = "Vac"
        if 'Alyona' in hello[9]:
            ws['K11'] = "Vac"
    if 'AlyonaVacSat' in val:
        if 'Alyona' in hello[0]:
            ws['L2'] = "Vac"
        if 'Alyona' in hello[1]:
            ws['L3'] = "Vac"
        if 'Alyona' in hello[2]:
            ws['L4'] = "Vac"
        if 'Alyona' in hello[3]:
            ws['L5'] = "Vac"
        if 'Alyona' in hello[4]:
            ws['L6'] = "Vac"
        if 'Alyona' in hello[5]:
            ws['L7'] = "Vac"
        if 'Alyona' in hello[6]:
            ws['L8'] = "Vac"
        if 'Alyona' in hello[7]:
            ws['L9'] = "Vac"
        if 'Alyona' in hello[8]:
            ws['L10'] = "Vac"
        if 'Alyona' in hello[9]:
            ws['L11'] = "Vac"
    if 'AlyonaVacAll' in val:
        if 'Alyona' in hello[0]:
            ws['A2'] = "Alyona-Vac"
            ws['B2'] = "Alyona-Vac"
            ws['C2'] = "Alyona-Vac"
            ws['D2'] = "Alyona-Vac"
            ws['E2'] = "Alyona-Vac"
            ws['F2'] = "Alyona-Vac"
            ws['G2'] = "Alyona-Vac"
            ws['H2'] = "Alyona-Vac"
            ws['I2'] = "Alyona-Vac"
            ws['J2'] = "Alyona-Vac"
            ws['K2'] = "Alyona-Vac"
            ws['L2'] = "Alyona-Vac"
        if 'Alyona' in hello[1]:
            ws['A3'] = "Alyona-Vac"
            ws['B3'] = "Alyona-Vac"
            ws['C3'] = "Alyona-Vac"
            ws['D3'] = "Alyona-Vac"
            ws['E3'] = "Alyona-Vac"
            ws['F3'] = "Alyona-Vac"
            ws['G3'] = "Alyona-Vac"
            ws['H3'] = "Alyona-Vac"
            ws['I3'] = "Alyona-Vac"
            ws['J3'] = "Alyona-Vac"
            ws['K3'] = "Alyona-Vac"
            ws['L3'] = "Alyona-Vac"
        if 'Alyona' in hello[2]:
            ws['A4'] = "Alyona-Vac"
            ws['B4'] = "Alyona-Vac"
            ws['C4'] = "Alyona-Vac"
            ws['D4'] = "Alyona-Vac"
            ws['E4'] = "Alyona-Vac"
            ws['F4'] = "Alyona-Vac"
            ws['G4'] = "Alyona-Vac"
            ws['H4'] = "Alyona-Vac"
            ws['I4'] = "Alyona-Vac"
            ws['J4'] = "Alyona-Vac"
            ws['K4'] = "Alyona-Vac"
            ws['L4'] = "Alyona-Vac"
        if 'Alyona' in hello[3]:
            ws['A5'] = "Alyona-Vac"
            ws['B5'] = "Alyona-Vac"
            ws['C5'] = "Alyona-Vac"
            ws['D5'] = "Alyona-Vac"
            ws['E5'] = "Alyona-Vac"
            ws['F5'] = "Alyona-Vac"
            ws['G5'] = "Alyona-Vac"
            ws['H5'] = "Alyona-Vac"
            ws['I5'] = "Alyona-Vac"
            ws['J5'] = "Alyona-Vac"
            ws['K5'] = "Alyona-Vac"
            ws['L5'] = "Alyona-Vac"
        if 'Alyona' in hello[4]:
            ws['A6'] = "Alyona-Vac"
            ws['B6'] = "Alyona-Vac"
            ws['C6'] = "Alyona-Vac"
            ws['D6'] = "Alyona-Vac"
            ws['E6'] = "Alyona-Vac"
            ws['F6'] = "Alyona-Vac"
            ws['G6'] = "Alyona-Vac"
            ws['H6'] = "Alyona-Vac"
            ws['I6'] = "Alyona-Vac"
            ws['J6'] = "Alyona-Vac"
            ws['K6'] = "Alyona-Vac"
            ws['L6'] = "Alyona-Vac"
        if 'Alyona' in hello[5]:
            ws['A7'] = "Alyona-Vac"
            ws['B7'] = "Alyona-Vac"
            ws['C7'] = "Alyona-Vac"
            ws['D7'] = "Alyona-Vac"
            ws['E7'] = "Alyona-Vac"
            ws['F7'] = "Alyona-Vac"
            ws['G7'] = "Alyona-Vac"
            ws['H7'] = "Alyona-Vac"
            ws['I7'] = "Alyona-Vac"
            ws['J7'] = "Alyona-Vac"
            ws['K7'] = "Alyona-Vac"
            ws['L7'] = "Alyona-Vac"
        if 'Alyona' in hello[6]:
            ws['A8'] = "Alyona-Vac"
            ws['B8'] = "Alyona-Vac"
            ws['C8'] = "Alyona-Vac"
            ws['D8'] = "Alyona-Vac"
            ws['E8'] = "Alyona-Vac"
            ws['F8'] = "Alyona-Vac"
            ws['G8'] = "Alyona-Vac"
            ws['H8'] = "Alyona-Vac"
            ws['I8'] = "Alyona-Vac"
            ws['J8'] = "Alyona-Vac"
            ws['K8'] = "Alyona-Vac"
            ws['L8'] = "Alyona-Vac"
        if 'Alyona' in hello[7]:
            ws['A9'] = "Alyona-Vac"
            ws['B9'] = "Alyona-Vac"
            ws['C9'] = "Alyona-Vac"
            ws['D9'] = "Alyona-Vac"
            ws['E9'] = "Alyona-Vac"
            ws['F9'] = "Alyona-Vac"
            ws['G9'] = "Alyona-Vac"
            ws['H9'] = "Alyona-Vac"
            ws['I9'] = "Alyona-Vac"
            ws['J9'] = "Alyona-Vac"
            ws['K9'] = "Alyona-Vac"
            ws['L9'] = "Alyona-Vac"
        if 'Alyona' in hello[8]:
            ws['A9'] = "Alyona-Vac"
            ws['B9'] = "Alyona-Vac"
            ws['C9'] = "Alyona-Vac"
            ws['D9'] = "Alyona-Vac"
            ws['E9'] = "Alyona-Vac"
            ws['F9'] = "Alyona-Vac"
            ws['G9'] = "Alyona-Vac"
            ws['H9'] = "Alyona-Vac"
            ws['I9'] = "Alyona-Vac"
            ws['J9'] = "Alyona-Vac"
            ws['K9'] = "Alyona-Vac"
            ws['L9'] = "Alyona-Vac"
        if 'Alyona' in hello[9]:
            ws['A10'] = "Alyona-Vac"
            ws['B10'] = "Alyona-Vac"
            ws['C10'] = "Alyona-Vac"
            ws['D10'] = "Alyona-Vac"
            ws['E10'] = "Alyona-Vac"
            ws['F10'] = "Alyona-Vac"
            ws['G10'] = "Alyona-Vac"
            ws['H10'] = "Alyona-Vac"
            ws['I10'] = "Alyona-Vac"
            ws['J10'] = "Alyona-Vac"
            ws['K10'] = "Alyona-Vac"
            ws['L10'] = "Alyona-Vac"
    if 'OfirVacSun' in val:
        if 'Ofir' in hello[0]:
            ws['A2'] = "Vac"
            ws['B2'] = "Vac"
        if 'Ofir' in hello[1]:
            ws['A3'] = "Vac"
            ws['B3'] = "Vac"
        if 'Ofir' in hello[2]:
            ws['A4'] = "Vac"
            ws['B4'] = "Vac"
        if 'Ofir' in hello[3]:
            ws['A5'] = "Vac"
            ws['B5'] = "Vac"
        if 'Ofir' in hello[4]:
            ws['A6'] = "Vac"
            ws['B6'] = "Vac"
        if 'Ofir' in hello[5]:
            ws['A7'] = "Vac"
            ws['B7'] = "Vac"
        if 'Ofir' in hello[6]:
            ws['A8'] = "Vac"
            ws['B8'] = "Vac"
        if 'Ofir' in hello[7]:
            ws['A9'] = "Vac"
            ws['B9'] = "Vac"
        if 'Ofir' in hello[8]:
            ws['A10'] = "Vac"
            ws['B10'] = "Vac"
        if 'Ofir' in hello[9]:
            ws['A11'] = "Vac"
            ws['B11'] = "Vac"
    if 'OfirVacMon' in val:
        if 'Ofir' in hello[0]:
            ws['C2'] = "Vac"
            ws['D2'] = "Vac"
        if 'Ofir' in hello[1]:
            ws['C3'] = "Vac"
            ws['D3'] = "Vac"
        if 'Ofir' in hello[2]:
            ws['C4'] = "Vac"
            ws['D4'] = "Vac"
        if 'Ofir' in hello[3]:
            ws['C5'] = "Vac"
            ws['D5'] = "Vac"
        if 'Ofir' in hello[4]:
            ws['C6'] = "Vac"
            ws['D6'] = "Vac"
        if 'Ofir' in hello[5]:
            ws['C7'] = "Vac"
            ws['D7'] = "Vac"
        if 'Ofir' in hello[6]:
            ws['C8'] = "Vac"
            ws['D8'] = "Vac"
        if 'Ofir' in hello[7]:
            ws['C9'] = "Vac"
            ws['D9'] = "Vac"
        if 'Ofir' in hello[8]:
            ws['C10'] = "Vac"
            ws['D10'] = "Vac"
        if 'Ofir' in hello[9]:
            ws['C11'] = "Vac"
            ws['D11'] = "Vac"
    if 'OfirVacTue' in val:
        if 'Ofir' in hello[0]:
            ws['E2'] = "Vac"
            ws['F2'] = "Vac"
        if 'Ofir' in hello[1]:
            ws['E3'] = "Vac"
            ws['F3'] = "Vac"
        if 'Ofir' in hello[2]:
            ws['E4'] = "Vac"
            ws['F4'] = "Vac"
        if 'Ofir' in hello[3]:
            ws['E5'] = "Vac"
            ws['F5'] = "Vac"
        if 'Ofir' in hello[4]:
            ws['E6'] = "Vac"
            ws['F6'] = "Vac"
        if 'Ofir' in hello[5]:
            ws['E7'] = "Vac"
            ws['F7'] = "Vac"
        if 'Ofir' in hello[6]:
            ws['E8'] = "Vac"
            ws['F8'] = "Vac"
        if 'Ofir' in hello[7]:
            ws['E9'] = "Vac"
            ws['F9'] = "Vac"
        if 'Ofir' in hello[8]:
            ws['E10'] = "Vac"
            ws['F10'] = "Vac"
        if 'Ofir' in hello[9]:
            ws['E11'] = "Vac"
            ws['F11'] = "Vac"
    if 'OfirVacWed' in val:
        if 'Ofir' in hello[0]:
            ws['G2'] = "Vac"
            ws['H2'] = "Vac"
        if 'Ofir' in hello[1]:
            ws['G3'] = "Vac"
            ws['H3'] = "Vac"
        if 'Ofir' in hello[2]:
            ws['G4'] = "Vac"
            ws['H4'] = "Vac"
        if 'Ofir' in hello[3]:
            ws['G5'] = "Vac"
            ws['H5'] = "Vac"
        if 'Ofir' in hello[4]:
            ws['G6'] = "Vac"
            ws['H6'] = "Vac"
        if 'Ofir' in hello[5]:
            ws['G7'] = "Vac"
            ws['H7'] = "Vac"
        if 'Ofir' in hello[6]:
            ws['G8'] = "Vac"
            ws['H8'] = "Vac"
        if 'Ofir' in hello[7]:
            ws['G9'] = "Vac"
            ws['H9'] = "Vac"
        if 'Ofir' in hello[8]:
            ws['G10'] = "Vac"
            ws['H10'] = "Vac"
        if 'Ofir' in hello[9]:
            ws['G11'] = "Vac"
            ws['H11'] = "Vac"
    if 'OfirVacThu' in val:
        if 'Ofir' in hello[0]:
            ws['I2'] = "Vac"
            ws['j2'] = "Vac"
        if 'Ofir' in hello[1]:
            ws['I3'] = "Vac"
            ws['j3'] = "Vac"
        if 'Ofir' in hello[2]:
            ws['I4'] = "Vac"
            ws['j4'] = "Vac"
        if 'Ofir' in hello[3]:
            ws['I5'] = "Vac"
            ws['j5'] = "Vac"
        if 'Ofir' in hello[4]:
            ws['I6'] = "Vac"
            ws['j6'] = "Vac"
        if 'Ofir' in hello[5]:
            ws['I7'] = "Vac"
            ws['j7'] = "Vac"
        if 'Ofir' in hello[6]:
            ws['I8'] = "Vac"
            ws['j8'] = "Vac"
        if 'Ofir' in hello[7]:
            ws['I9'] = "Vac"
            ws['j9'] = "Vac"
        if 'Ofir' in hello[8]:
            ws['I10'] = "Vac"
            ws['j10'] = "Vac"
        if 'Ofir' in hello[9]:
            ws['I11'] = "Vac"
            ws['j11'] = "Vac"
    if 'OfirVacFri' in val:
        if 'Ofir' in hello[0]:
            ws['K2'] = "Vac"
        if 'Ofir' in hello[1]:
            ws['K3'] = "Vac"
        if 'Ofir' in hello[2]:
            ws['K4'] = "Vac"
        if 'Ofir' in hello[3]:
            ws['K5'] = "Vac"
        if 'Ofir' in hello[4]:
            ws['K6'] = "Vac"
        if 'Ofir' in hello[5]:
            ws['K7'] = "Vac"
        if 'Ofir' in hello[6]:
            ws['K8'] = "Vac"
        if 'Ofir' in hello[7]:
            ws['K9'] = "Vac"
        if 'Ofir' in hello[8]:
            ws['K10'] = "Vac"
        if 'Ofir' in hello[9]:
            ws['K11'] = "Vac"
    if 'OfirVacSat' in val:
        if 'Ofir' in hello[0]:
            ws['L2'] = "Vac"
        if 'Ofir' in hello[1]:
            ws['L3'] = "Vac"
        if 'Ofir' in hello[2]:
            ws['L4'] = "Vac"
        if 'Ofir' in hello[3]:
            ws['L5'] = "Vac"
        if 'Ofir' in hello[4]:
            ws['L6'] = "Vac"
        if 'Ofir' in hello[5]:
            ws['L7'] = "Vac"
        if 'Ofir' in hello[6]:
            ws['L8'] = "Vac"
        if 'Ofir' in hello[7]:
            ws['L9'] = "Vac"
        if 'Ofir' in hello[8]:
            ws['L10'] = "Vac"
        if 'Ofir' in hello[9]:
            ws['L11'] = "Vac"
    if 'OfirVacAll' in val:
        if 'Ofir' in hello[0]:
            ws['A2'] = "Ofir-Vac"
            ws['B2'] = "Ofir-Vac"
            ws['C2'] = "Ofir-Vac"
            ws['D2'] = "Ofir-Vac"
            ws['E2'] = "Ofir-Vac"
            ws['F2'] = "Ofir-Vac"
            ws['G2'] = "Ofir-Vac"
            ws['H2'] = "Ofir-Vac"
            ws['I2'] = "Ofir-Vac"
            ws['J2'] = "Ofir-Vac"
            ws['K2'] = "Ofir-Vac"
            ws['L2'] = "Ofir-Vac"
        if 'Ofir' in hello[1]:
            ws['A3'] = "Ofir-Vac"
            ws['B3'] = "Ofir-Vac"
            ws['C3'] = "Ofir-Vac"
            ws['D3'] = "Ofir-Vac"
            ws['E3'] = "Ofir-Vac"
            ws['F3'] = "Ofir-Vac"
            ws['G3'] = "Ofir-Vac"
            ws['H3'] = "Ofir-Vac"
            ws['I3'] = "Ofir-Vac"
            ws['J3'] = "Ofir-Vac"
            ws['K3'] = "Ofir-Vac"
            ws['L3'] = "Ofir-Vac"
        if 'Ofir' in hello[2]:
            ws['A4'] = "Ofir-Vac"
            ws['B4'] = "Ofir-Vac"
            ws['C4'] = "Ofir-Vac"
            ws['D4'] = "Ofir-Vac"
            ws['E4'] = "Ofir-Vac"
            ws['F4'] = "Ofir-Vac"
            ws['G4'] = "Ofir-Vac"
            ws['H4'] = "Ofir-Vac"
            ws['I4'] = "Ofir-Vac"
            ws['J4'] = "Ofir-Vac"
            ws['K4'] = "Ofir-Vac"
            ws['L4'] = "Ofir-Vac"
        if 'Ofir' in hello[3]:
            ws['A5'] = "Ofir-Vac"
            ws['B5'] = "Ofir-Vac"
            ws['C5'] = "Ofir-Vac"
            ws['D5'] = "Ofir-Vac"
            ws['E5'] = "Ofir-Vac"
            ws['F5'] = "Ofir-Vac"
            ws['G5'] = "Ofir-Vac"
            ws['H5'] = "Ofir-Vac"
            ws['I5'] = "Ofir-Vac"
            ws['J5'] = "Ofir-Vac"
            ws['K5'] = "Ofir-Vac"
            ws['L5'] = "Ofir-Vac"
        if 'Ofir' in hello[4]:
            ws['A6'] = "Ofir-Vac"
            ws['B6'] = "Ofir-Vac"
            ws['C6'] = "Ofir-Vac"
            ws['D6'] = "Ofir-Vac"
            ws['E6'] = "Ofir-Vac"
            ws['F6'] = "Ofir-Vac"
            ws['G6'] = "Ofir-Vac"
            ws['H6'] = "Ofir-Vac"
            ws['I6'] = "Ofir-Vac"
            ws['J6'] = "Ofir-Vac"
            ws['K6'] = "Ofir-Vac"
            ws['L6'] = "Ofir-Vac"
        if 'Ofir' in hello[5]:
            ws['A7'] = "Ofir-Vac"
            ws['B7'] = "Ofir-Vac"
            ws['C7'] = "Ofir-Vac"
            ws['D7'] = "Ofir-Vac"
            ws['E7'] = "Ofir-Vac"
            ws['F7'] = "Ofir-Vac"
            ws['G7'] = "Ofir-Vac"
            ws['H7'] = "Ofir-Vac"
            ws['I7'] = "Ofir-Vac"
            ws['J7'] = "Ofir-Vac"
            ws['K7'] = "Ofir-Vac"
            ws['L7'] = "Ofir-Vac"
        if 'Ofir' in hello[6]:
            ws['A8'] = "Ofir-Vac"
            ws['B8'] = "Ofir-Vac"
            ws['C8'] = "Ofir-Vac"
            ws['D8'] = "Ofir-Vac"
            ws['E8'] = "Ofir-Vac"
            ws['F8'] = "Ofir-Vac"
            ws['G8'] = "Ofir-Vac"
            ws['H8'] = "Ofir-Vac"
            ws['I8'] = "Ofir-Vac"
            ws['J8'] = "Ofir-Vac"
            ws['K8'] = "Ofir-Vac"
            ws['L8'] = "Ofir-Vac"
        if 'Ofir' in hello[7]:
            ws['A9'] = "Ofir-Vac"
            ws['B9'] = "Ofir-Vac"
            ws['C9'] = "Ofir-Vac"
            ws['D9'] = "Ofir-Vac"
            ws['E9'] = "Ofir-Vac"
            ws['F9'] = "Ofir-Vac"
            ws['G9'] = "Ofir-Vac"
            ws['H9'] = "Ofir-Vac"
            ws['I9'] = "Ofir-Vac"
            ws['J9'] = "Ofir-Vac"
            ws['K9'] = "Ofir-Vac"
            ws['L9'] = "Ofir-Vac"
        if 'Ofir' in hello[8]:
            ws['A9'] = "Ofir-Vac"
            ws['B9'] = "Ofir-Vac"
            ws['C9'] = "Ofir-Vac"
            ws['D9'] = "Ofir-Vac"
            ws['E9'] = "Ofir-Vac"
            ws['F9'] = "Ofir-Vac"
            ws['G9'] = "Ofir-Vac"
            ws['H9'] = "Ofir-Vac"
            ws['I9'] = "Ofir-Vac"
            ws['J9'] = "Ofir-Vac"
            ws['K9'] = "Ofir-Vac"
            ws['L9'] = "Ofir-Vac"
        if 'Ofir' in hello[9]:
            ws['A10'] = "Ofir-Vac"
            ws['B10'] = "Ofir-Vac"
            ws['C10'] = "Ofir-Vac"
            ws['D10'] = "Ofir-Vac"
            ws['E10'] = "Ofir-Vac"
            ws['F10'] = "Ofir-Vac"
            ws['G10'] = "Ofir-Vac"
            ws['H10'] = "Ofir-Vac"
            ws['I10'] = "Ofir-Vac"
            ws['J10'] = "Ofir-Vac"
            ws['K10'] = "Ofir-Vac"
            ws['L10'] = "Ofir-Vac"
    if 'YairVacSun' in val:
        if 'Yair' in hello[0]:
            ws['A2'] = "Vac"
            ws['B2'] = "Vac"
        if 'Yair' in hello[1]:
            ws['A3'] = "Vac"
            ws['B3'] = "Vac"
        if 'Yair' in hello[2]:
            ws['A4'] = "Vac"
            ws['B4'] = "Vac"
        if 'Yair' in hello[3]:
            ws['A5'] = "Vac"
            ws['B5'] = "Vac"
        if 'Yair' in hello[4]:
            ws['A6'] = "Vac"
            ws['B6'] = "Vac"
        if 'Yair' in hello[5]:
            ws['A7'] = "Vac"
            ws['B7'] = "Vac"
        if 'Yair' in hello[6]:
            ws['A8'] = "Vac"
            ws['B8'] = "Vac"
        if 'Yair' in hello[7]:
            ws['A9'] = "Vac"
            ws['B9'] = "Vac"
        if 'Yair' in hello[8]:
            ws['A10'] = "Vac"
            ws['B10'] = "Vac"
        if 'Yair' in hello[9]:
            ws['A11'] = "Vac"
            ws['B11'] = "Vac"
    if 'YairVacMon' in val:
        if 'Yair' in hello[0]:
            ws['C2'] = "Vac"
            ws['D2'] = "Vac"
        if 'Yair' in hello[1]:
            ws['C3'] = "Vac"
            ws['D3'] = "Vac"
        if 'Yair' in hello[2]:
            ws['C4'] = "Vac"
            ws['D4'] = "Vac"
        if 'Yair' in hello[3]:
            ws['C5'] = "Vac"
            ws['D5'] = "Vac"
        if 'Yair' in hello[4]:
            ws['C6'] = "Vac"
            ws['D6'] = "Vac"
        if 'Yair' in hello[5]:
            ws['C7'] = "Vac"
            ws['D7'] = "Vac"
        if 'Yair' in hello[6]:
            ws['C8'] = "Vac"
            ws['D8'] = "Vac"
        if 'Yair' in hello[7]:
            ws['C9'] = "Vac"
            ws['D9'] = "Vac"
        if 'Yair' in hello[8]:
            ws['C10'] = "Vac"
            ws['D10'] = "Vac"
        if 'Yair' in hello[9]:
            ws['C11'] = "Vac"
            ws['D11'] = "Vac"
    if 'YairVacTue' in val:
        if 'Yair' in hello[0]:
            ws['E2'] = "Vac"
            ws['F2'] = "Vac"
        if 'Yair' in hello[1]:
            ws['E3'] = "Vac"
            ws['F3'] = "Vac"
        if 'Yair' in hello[2]:
            ws['E4'] = "Vac"
            ws['F4'] = "Vac"
        if 'Yair' in hello[3]:
            ws['E5'] = "Vac"
            ws['F5'] = "Vac"
        if 'Yair' in hello[4]:
            ws['E6'] = "Vac"
            ws['F6'] = "Vac"
        if 'Yair' in hello[5]:
            ws['E7'] = "Vac"
            ws['F7'] = "Vac"
        if 'Yair' in hello[6]:
            ws['E8'] = "Vac"
            ws['F8'] = "Vac"
        if 'Yair' in hello[7]:
            ws['E9'] = "Vac"
            ws['F9'] = "Vac"
        if 'Yair' in hello[8]:
            ws['E10'] = "Vac"
            ws['F10'] = "Vac"
        if 'Yair' in hello[9]:
            ws['E11'] = "Vac"
            ws['F11'] = "Vac"
    if 'YairVacWed' in val:
        if 'Yair' in hello[0]:
            ws['G2'] = "Vac"
            ws['H2'] = "Vac"
        if 'Yair' in hello[1]:
            ws['G3'] = "Vac"
            ws['H3'] = "Vac"
        if 'Yair' in hello[2]:
            ws['G4'] = "Vac"
            ws['H4'] = "Vac"
        if 'Yair' in hello[3]:
            ws['G5'] = "Vac"
            ws['H5'] = "Vac"
        if 'Yair' in hello[4]:
            ws['G6'] = "Vac"
            ws['H6'] = "Vac"
        if 'Yair' in hello[5]:
            ws['G7'] = "Vac"
            ws['H7'] = "Vac"
        if 'Yair' in hello[6]:
            ws['G8'] = "Vac"
            ws['H8'] = "Vac"
        if 'Yair' in hello[7]:
            ws['G9'] = "Vac"
            ws['H9'] = "Vac"
        if 'Yair' in hello[8]:
            ws['G10'] = "Vac"
            ws['H10'] = "Vac"
        if 'Yair' in hello[9]:
            ws['G11'] = "Vac"
            ws['H11'] = "Vac"
    if 'YairVacThu' in val:
        if 'Yair' in hello[0]:
            ws['I2'] = "Vac"
            ws['j2'] = "Vac"
        if 'Yair' in hello[1]:
            ws['I3'] = "Vac"
            ws['j3'] = "Vac"
        if 'Yair' in hello[2]:
            ws['I4'] = "Vac"
            ws['j4'] = "Vac"
        if 'Yair' in hello[3]:
            ws['I5'] = "Vac"
            ws['j5'] = "Vac"
        if 'Yair' in hello[4]:
            ws['I6'] = "Vac"
            ws['j6'] = "Vac"
        if 'Yair' in hello[5]:
            ws['I7'] = "Vac"
            ws['j7'] = "Vac"
        if 'Yair' in hello[6]:
            ws['I8'] = "Vac"
            ws['j8'] = "Vac"
        if 'Yair' in hello[7]:
            ws['I9'] = "Vac"
            ws['j9'] = "Vac"
        if 'Yair' in hello[8]:
            ws['I10'] = "Vac"
            ws['j10'] = "Vac"
        if 'Yair' in hello[9]:
            ws['I11'] = "Vac"
            ws['j11'] = "Vac"
    if 'YairVacFri' in val:
        if 'Yair' in hello[0]:
            ws['K2'] = "Vac"
        if 'Yair' in hello[1]:
            ws['K3'] = "Vac"
        if 'Yair' in hello[2]:
            ws['K4'] = "Vac"
        if 'Yair' in hello[3]:
            ws['K5'] = "Vac"
        if 'Yair' in hello[4]:
            ws['K6'] = "Vac"
        if 'Yair' in hello[5]:
            ws['K7'] = "Vac"
        if 'Yair' in hello[6]:
            ws['K8'] = "Vac"
        if 'Yair' in hello[7]:
            ws['K9'] = "Vac"
        if 'Yair' in hello[8]:
            ws['K10'] = "Vac"
        if 'Yair' in hello[9]:
            ws['K11'] = "Vac"
    if 'YairVacSat' in val:
        if 'Yair' in hello[0]:
            ws['L2'] = "Vac"
        if 'Yair' in hello[1]:
            ws['L3'] = "Vac"
        if 'Yair' in hello[2]:
            ws['L4'] = "Vac"
        if 'Yair' in hello[3]:
            ws['L5'] = "Vac"
        if 'Yair' in hello[4]:
            ws['L6'] = "Vac"
        if 'Yair' in hello[5]:
            ws['L7'] = "Vac"
        if 'Yair' in hello[6]:
            ws['L8'] = "Vac"
        if 'Yair' in hello[7]:
            ws['L9'] = "Vac"
        if 'Yair' in hello[8]:
            ws['L10'] = "Vac"
        if 'Yair' in hello[9]:
            ws['L11'] = "Vac"
    if 'YairVacAll' in val:
        if 'Yair' in hello[0]:
            ws['A2'] = "Yair-Vac"
            ws['B2'] = "Yair-Vac"
            ws['C2'] = "Yair-Vac"
            ws['D2'] = "Yair-Vac"
            ws['E2'] = "Yair-Vac"
            ws['F2'] = "Yair-Vac"
            ws['G2'] = "Yair-Vac"
            ws['H2'] = "Yair-Vac"
            ws['I2'] = "Yair-Vac"
            ws['J2'] = "Yair-Vac"
            ws['K2'] = "Yair-Vac"
            ws['L2'] = "Yair-Vac"
        if 'Yair' in hello[1]:
            ws['A3'] = "Yair-Vac"
            ws['B3'] = "Yair-Vac"
            ws['C3'] = "Yair-Vac"
            ws['D3'] = "Yair-Vac"
            ws['E3'] = "Yair-Vac"
            ws['F3'] = "Yair-Vac"
            ws['G3'] = "Yair-Vac"
            ws['H3'] = "Yair-Vac"
            ws['I3'] = "Yair-Vac"
            ws['J3'] = "Yair-Vac"
            ws['K3'] = "Yair-Vac"
            ws['L3'] = "Yair-Vac"
        if 'Yair' in hello[2]:
            ws['A4'] = "Yair-Vac"
            ws['B4'] = "Yair-Vac"
            ws['C4'] = "Yair-Vac"
            ws['D4'] = "Yair-Vac"
            ws['E4'] = "Yair-Vac"
            ws['F4'] = "Yair-Vac"
            ws['G4'] = "Yair-Vac"
            ws['H4'] = "Yair-Vac"
            ws['I4'] = "Yair-Vac"
            ws['J4'] = "Yair-Vac"
            ws['K4'] = "Yair-Vac"
            ws['L4'] = "Yair-Vac"
        if 'Yair' in hello[3]:
            ws['A5'] = "Yair-Vac"
            ws['B5'] = "Yair-Vac"
            ws['C5'] = "Yair-Vac"
            ws['D5'] = "Yair-Vac"
            ws['E5'] = "Yair-Vac"
            ws['F5'] = "Yair-Vac"
            ws['G5'] = "Yair-Vac"
            ws['H5'] = "Yair-Vac"
            ws['I5'] = "Yair-Vac"
            ws['J5'] = "Yair-Vac"
            ws['K5'] = "Yair-Vac"
            ws['L5'] = "Yair-Vac"
        if 'Yair' in hello[4]:
            ws['A6'] = "Yair-Vac"
            ws['B6'] = "Yair-Vac"
            ws['C6'] = "Yair-Vac"
            ws['D6'] = "Yair-Vac"
            ws['E6'] = "Yair-Vac"
            ws['F6'] = "Yair-Vac"
            ws['G6'] = "Yair-Vac"
            ws['H6'] = "Yair-Vac"
            ws['I6'] = "Yair-Vac"
            ws['J6'] = "Yair-Vac"
            ws['K6'] = "Yair-Vac"
            ws['L6'] = "Yair-Vac"
        if 'Yair' in hello[5]:
            ws['A7'] = "Yair-Vac"
            ws['B7'] = "Yair-Vac"
            ws['C7'] = "Yair-Vac"
            ws['D7'] = "Yair-Vac"
            ws['E7'] = "Yair-Vac"
            ws['F7'] = "Yair-Vac"
            ws['G7'] = "Yair-Vac"
            ws['H7'] = "Yair-Vac"
            ws['I7'] = "Yair-Vac"
            ws['J7'] = "Yair-Vac"
            ws['K7'] = "Yair-Vac"
            ws['L7'] = "Yair-Vac"
        if 'Yair' in hello[6]:
            ws['A8'] = "Yair-Vac"
            ws['B8'] = "Yair-Vac"
            ws['C8'] = "Yair-Vac"
            ws['D8'] = "Yair-Vac"
            ws['E8'] = "Yair-Vac"
            ws['F8'] = "Yair-Vac"
            ws['G8'] = "Yair-Vac"
            ws['H8'] = "Yair-Vac"
            ws['I8'] = "Yair-Vac"
            ws['J8'] = "Yair-Vac"
            ws['K8'] = "Yair-Vac"
            ws['L8'] = "Yair-Vac"
        if 'Yair' in hello[7]:
            ws['A9'] = "Yair-Vac"
            ws['B9'] = "Yair-Vac"
            ws['C9'] = "Yair-Vac"
            ws['D9'] = "Yair-Vac"
            ws['E9'] = "Yair-Vac"
            ws['F9'] = "Yair-Vac"
            ws['G9'] = "Yair-Vac"
            ws['H9'] = "Yair-Vac"
            ws['I9'] = "Yair-Vac"
            ws['J9'] = "Yair-Vac"
            ws['K9'] = "Yair-Vac"
            ws['L9'] = "Yair-Vac"
        if 'Yair' in hello[8]:
            ws['A9'] = "Yair-Vac"
            ws['B9'] = "Yair-Vac"
            ws['C9'] = "Yair-Vac"
            ws['D9'] = "Yair-Vac"
            ws['E9'] = "Yair-Vac"
            ws['F9'] = "Yair-Vac"
            ws['G9'] = "Yair-Vac"
            ws['H9'] = "Yair-Vac"
            ws['I9'] = "Yair-Vac"
            ws['J9'] = "Yair-Vac"
            ws['K9'] = "Yair-Vac"
            ws['L9'] = "Yair-Vac"
        if 'Yair' in hello[9]:
            ws['A10'] = "Yair-Vac"
            ws['B10'] = "Yair-Vac"
            ws['C10'] = "Yair-Vac"
            ws['D10'] = "Yair-Vac"
            ws['E10'] = "Yair-Vac"
            ws['F10'] = "Yair-Vac"
            ws['G10'] = "Yair-Vac"
            ws['H10'] = "Yair-Vac"
            ws['I10'] = "Yair-Vac"
            ws['J10'] = "Yair-Vac"
            ws['K10'] = "Yair-Vac"
            ws['L10'] = "Yair-Vac"
    if 'RanVacSun' in val:
        if 'Ran' in hello[0]:
            ws['A2'] = "Vac"
            ws['B2'] = "Vac"
        if 'Ran' in hello[1]:
            ws['A3'] = "Vac"
            ws['B3'] = "Vac"
        if 'Ran' in hello[2]:
            ws['A4'] = "Vac"
            ws['B4'] = "Vac"
        if 'Ran' in hello[3]:
            ws['A5'] = "Vac"
            ws['B5'] = "Vac"
        if 'Ran' in hello[4]:
            ws['A6'] = "Vac"
            ws['B6'] = "Vac"
        if 'Ran' in hello[5]:
            ws['A7'] = "Vac"
            ws['B7'] = "Vac"
        if 'Ran' in hello[6]:
            ws['A8'] = "Vac"
            ws['B8'] = "Vac"
        if 'Ran' in hello[7]:
            ws['A9'] = "Vac"
            ws['B9'] = "Vac"
        if 'Ran' in hello[8]:
            ws['A10'] = "Vac"
            ws['B10'] = "Vac"
        if 'Ran' in hello[9]:
            ws['A11'] = "Vac"
            ws['B11'] = "Vac"
    if 'RanVacMon' in val:
        if 'Ran' in hello[0]:
            ws['C2'] = "Vac"
            ws['D2'] = "Vac"
        if 'Ran' in hello[1]:
            ws['C3'] = "Vac"
            ws['D3'] = "Vac"
        if 'Ran' in hello[2]:
            ws['C4'] = "Vac"
            ws['D4'] = "Vac"
        if 'Ran' in hello[3]:
            ws['C5'] = "Vac"
            ws['D5'] = "Vac"
        if 'Ran' in hello[4]:
            ws['C6'] = "Vac"
            ws['D6'] = "Vac"
        if 'Ran' in hello[5]:
            ws['C7'] = "Vac"
            ws['D7'] = "Vac"
        if 'Ran' in hello[6]:
            ws['C8'] = "Vac"
            ws['D8'] = "Vac"
        if 'Ran' in hello[7]:
            ws['C9'] = "Vac"
            ws['D9'] = "Vac"
        if 'Ran' in hello[8]:
            ws['C10'] = "Vac"
            ws['D10'] = "Vac"
        if 'Ran' in hello[9]:
            ws['C11'] = "Vac"
            ws['D11'] = "Vac"
    if 'RanVacTue' in val:
        if 'Ran' in hello[0]:
            ws['E2'] = "Vac"
            ws['F2'] = "Vac"
        if 'Ran' in hello[1]:
            ws['E3'] = "Vac"
            ws['F3'] = "Vac"
        if 'Ran' in hello[2]:
            ws['E4'] = "Vac"
            ws['F4'] = "Vac"
        if 'Ran' in hello[3]:
            ws['E5'] = "Vac"
            ws['F5'] = "Vac"
        if 'Ran' in hello[4]:
            ws['E6'] = "Vac"
            ws['F6'] = "Vac"
        if 'Ran' in hello[5]:
            ws['E7'] = "Vac"
            ws['F7'] = "Vac"
        if 'Ran' in hello[6]:
            ws['E8'] = "Vac"
            ws['F8'] = "Vac"
        if 'Ran' in hello[7]:
            ws['E9'] = "Vac"
            ws['F9'] = "Vac"
        if 'Ran' in hello[8]:
            ws['E10'] = "Vac"
            ws['F10'] = "Vac"
        if 'Ran' in hello[9]:
            ws['E11'] = "Vac"
            ws['F11'] = "Vac"
    if 'RanVacWed' in val:
        if 'Ran' in hello[0]:
            ws['G2'] = "Vac"
            ws['H2'] = "Vac"
        if 'Ran' in hello[1]:
            ws['G3'] = "Vac"
            ws['H3'] = "Vac"
        if 'Ran' in hello[2]:
            ws['G4'] = "Vac"
            ws['H4'] = "Vac"
        if 'Ran' in hello[3]:
            ws['G5'] = "Vac"
            ws['H5'] = "Vac"
        if 'Ran' in hello[4]:
            ws['G6'] = "Vac"
            ws['H6'] = "Vac"
        if 'Ran' in hello[5]:
            ws['G7'] = "Vac"
            ws['H7'] = "Vac"
        if 'Ran' in hello[6]:
            ws['G8'] = "Vac"
            ws['H8'] = "Vac"
        if 'Ran' in hello[7]:
            ws['G9'] = "Vac"
            ws['H9'] = "Vac"
        if 'Ran' in hello[8]:
            ws['G10'] = "Vac"
            ws['H10'] = "Vac"
        if 'Ran' in hello[9]:
            ws['G11'] = "Vac"
            ws['H11'] = "Vac"
    if 'RanVacThu' in val:
        if 'Ran' in hello[0]:
            ws['I2'] = "Vac"
            ws['j2'] = "Vac"
        if 'Ran' in hello[1]:
            ws['I3'] = "Vac"
            ws['j3'] = "Vac"
        if 'Ran' in hello[2]:
            ws['I4'] = "Vac"
            ws['j4'] = "Vac"
        if 'Ran' in hello[3]:
            ws['I5'] = "Vac"
            ws['j5'] = "Vac"
        if 'Ran' in hello[4]:
            ws['I6'] = "Vac"
            ws['j6'] = "Vac"
        if 'Ran' in hello[5]:
            ws['I7'] = "Vac"
            ws['j7'] = "Vac"
        if 'Ran' in hello[6]:
            ws['I8'] = "Vac"
            ws['j8'] = "Vac"
        if 'Ran' in hello[7]:
            ws['I9'] = "Vac"
            ws['j9'] = "Vac"
        if 'Ran' in hello[8]:
            ws['I10'] = "Vac"
            ws['j10'] = "Vac"
        if 'Ran' in hello[9]:
            ws['I11'] = "Vac"
            ws['j11'] = "Vac"
    if 'RanVacFri' in val:
        if 'Ran' in hello[0]:
            ws['K2'] = "Vac"
        if 'Ran' in hello[1]:
            ws['K3'] = "Vac"
        if 'Ran' in hello[2]:
            ws['K4'] = "Vac"
        if 'Ran' in hello[3]:
            ws['K5'] = "Vac"
        if 'Ran' in hello[4]:
            ws['K6'] = "Vac"
        if 'Ran' in hello[5]:
            ws['K7'] = "Vac"
        if 'Ran' in hello[6]:
            ws['K8'] = "Vac"
        if 'Ran' in hello[7]:
            ws['K9'] = "Vac"
        if 'Ran' in hello[8]:
            ws['K10'] = "Vac"
        if 'Ran' in hello[9]:
            ws['K11'] = "Vac"
    if 'RanVacSat' in val:
        if 'Ran' in hello[0]:
            ws['L2'] = "Vac"
        if 'Ran' in hello[1]:
            ws['L3'] = "Vac"
        if 'Ran' in hello[2]:
            ws['L4'] = "Vac"
        if 'Ran' in hello[3]:
            ws['L5'] = "Vac"
        if 'Ran' in hello[4]:
            ws['L6'] = "Vac"
        if 'Ran' in hello[5]:
            ws['L7'] = "Vac"
        if 'Ran' in hello[6]:
            ws['L8'] = "Vac"
        if 'Ran' in hello[7]:
            ws['L9'] = "Vac"
        if 'Ran' in hello[8]:
            ws['L10'] = "Vac"
        if 'Ran' in hello[9]:
            ws['L11'] = "Vac"
    if 'RanVacAll' in val:
        if 'Ran' in hello[0]:
            ws['A2'] = "Ran-Vac"
            ws['B2'] = "Ran-Vac"
            ws['C2'] = "Ran-Vac"
            ws['D2'] = "Ran-Vac"
            ws['E2'] = "Ran-Vac"
            ws['F2'] = "Ran-Vac"
            ws['G2'] = "Ran-Vac"
            ws['H2'] = "Ran-Vac"
            ws['I2'] = "Ran-Vac"
            ws['J2'] = "Ran-Vac"
            ws['K2'] = "Ran-Vac"
            ws['L2'] = "Ran-Vac"
        if 'Ran' in hello[1]:
            ws['A3'] = "Ran-Vac"
            ws['B3'] = "Ran-Vac"
            ws['C3'] = "Ran-Vac"
            ws['D3'] = "Ran-Vac"
            ws['E3'] = "Ran-Vac"
            ws['F3'] = "Ran-Vac"
            ws['G3'] = "Ran-Vac"
            ws['H3'] = "Ran-Vac"
            ws['I3'] = "Ran-Vac"
            ws['J3'] = "Ran-Vac"
            ws['K3'] = "Ran-Vac"
            ws['L3'] = "Ran-Vac"
        if 'Ran' in hello[2]:
            ws['A4'] = "Ran-Vac"
            ws['B4'] = "Ran-Vac"
            ws['C4'] = "Ran-Vac"
            ws['D4'] = "Ran-Vac"
            ws['E4'] = "Ran-Vac"
            ws['F4'] = "Ran-Vac"
            ws['G4'] = "Ran-Vac"
            ws['H4'] = "Ran-Vac"
            ws['I4'] = "Ran-Vac"
            ws['J4'] = "Ran-Vac"
            ws['K4'] = "Ran-Vac"
            ws['L4'] = "Ran-Vac"
        if 'Ran' in hello[3]:
            ws['A5'] = "Ran-Vac"
            ws['B5'] = "Ran-Vac"
            ws['C5'] = "Ran-Vac"
            ws['D5'] = "Ran-Vac"
            ws['E5'] = "Ran-Vac"
            ws['F5'] = "Ran-Vac"
            ws['G5'] = "Ran-Vac"
            ws['H5'] = "Ran-Vac"
            ws['I5'] = "Ran-Vac"
            ws['J5'] = "Ran-Vac"
            ws['K5'] = "Ran-Vac"
            ws['L5'] = "Ran-Vac"
        if 'Ran' in hello[4]:
            ws['A6'] = "Ran-Vac"
            ws['B6'] = "Ran-Vac"
            ws['C6'] = "Ran-Vac"
            ws['D6'] = "Ran-Vac"
            ws['E6'] = "Ran-Vac"
            ws['F6'] = "Ran-Vac"
            ws['G6'] = "Ran-Vac"
            ws['H6'] = "Ran-Vac"
            ws['I6'] = "Ran-Vac"
            ws['J6'] = "Ran-Vac"
            ws['K6'] = "Ran-Vac"
            ws['L6'] = "Ran-Vac"
        if 'Ran' in hello[5]:
            ws['A7'] = "Ran-Vac"
            ws['B7'] = "Ran-Vac"
            ws['C7'] = "Ran-Vac"
            ws['D7'] = "Ran-Vac"
            ws['E7'] = "Ran-Vac"
            ws['F7'] = "Ran-Vac"
            ws['G7'] = "Ran-Vac"
            ws['H7'] = "Ran-Vac"
            ws['I7'] = "Ran-Vac"
            ws['J7'] = "Ran-Vac"
            ws['K7'] = "Ran-Vac"
            ws['L7'] = "Ran-Vac"
        if 'Ran' in hello[6]:
            ws['A8'] = "Ran-Vac"
            ws['B8'] = "Ran-Vac"
            ws['C8'] = "Ran-Vac"
            ws['D8'] = "Ran-Vac"
            ws['E8'] = "Ran-Vac"
            ws['F8'] = "Ran-Vac"
            ws['G8'] = "Ran-Vac"
            ws['H8'] = "Ran-Vac"
            ws['I8'] = "Ran-Vac"
            ws['J8'] = "Ran-Vac"
            ws['K8'] = "Ran-Vac"
            ws['L8'] = "Ran-Vac"
        if 'Ran' in hello[7]:
            ws['A9'] = "Ran-Vac"
            ws['B9'] = "Ran-Vac"
            ws['C9'] = "Ran-Vac"
            ws['D9'] = "Ran-Vac"
            ws['E9'] = "Ran-Vac"
            ws['F9'] = "Ran-Vac"
            ws['G9'] = "Ran-Vac"
            ws['H9'] = "Ran-Vac"
            ws['I9'] = "Ran-Vac"
            ws['J9'] = "Ran-Vac"
            ws['K9'] = "Ran-Vac"
            ws['L9'] = "Ran-Vac"
        if 'Ran' in hello[8]:
            ws['A9'] = "Ran-Vac"
            ws['B9'] = "Ran-Vac"
            ws['C9'] = "Ran-Vac"
            ws['D9'] = "Ran-Vac"
            ws['E9'] = "Ran-Vac"
            ws['F9'] = "Ran-Vac"
            ws['G9'] = "Ran-Vac"
            ws['H9'] = "Ran-Vac"
            ws['I9'] = "Ran-Vac"
            ws['J9'] = "Ran-Vac"
            ws['K9'] = "Ran-Vac"
            ws['L9'] = "Ran-Vac"
        if 'Ran' in hello[9]:
            ws['A10'] = "Ran-Vac"
            ws['B10'] = "Ran-Vac"
            ws['C10'] = "Ran-Vac"
            ws['D10'] = "Ran-Vac"
            ws['E10'] = "Ran-Vac"
            ws['F10'] = "Ran-Vac"
            ws['G10'] = "Ran-Vac"
            ws['H10'] = "Ran-Vac"
            ws['I10'] = "Ran-Vac"
            ws['J10'] = "Ran-Vac"
            ws['K10'] = "Ran-Vac"
            ws['L10'] = "Ran-Vac"
    if 'AlexVacSun' in val:
        if 'Alex' in hello[0]:
            ws['A2'] = "Vac"
            ws['B2'] = "Vac"
        if 'Alex' in hello[1]:
            ws['A3'] = "Vac"
            ws['B3'] = "Vac"
        if 'Alex' in hello[2]:
            ws['A4'] = "Vac"
            ws['B4'] = "Vac"
        if 'Alex' in hello[3]:
            ws['A5'] = "Vac"
            ws['B5'] = "Vac"
        if 'Alex' in hello[4]:
            ws['A6'] = "Vac"
            ws['B6'] = "Vac"
        if 'Alex' in hello[5]:
            ws['A7'] = "Vac"
            ws['B7'] = "Vac"
        if 'Alex' in hello[6]:
            ws['A8'] = "Vac"
            ws['B8'] = "Vac"
        if 'Alex' in hello[7]:
            ws['A9'] = "Vac"
            ws['B9'] = "Vac"
        if 'Alex' in hello[8]:
            ws['A10'] = "Vac"
            ws['B10'] = "Vac"
        if 'Alex' in hello[9]:
            ws['A11'] = "Vac"
            ws['B11'] = "Vac"
    if 'AlexVacMon' in val:
        if 'Alex' in hello[0]:
            ws['C2'] = "Vac"
            ws['D2'] = "Vac"
        if 'Alex' in hello[1]:
            ws['C3'] = "Vac"
            ws['D3'] = "Vac"
        if 'Alex' in hello[2]:
            ws['C4'] = "Vac"
            ws['D4'] = "Vac"
        if 'Alex' in hello[3]:
            ws['C5'] = "Vac"
            ws['D5'] = "Vac"
        if 'Alex' in hello[4]:
            ws['C6'] = "Vac"
            ws['D6'] = "Vac"
        if 'Alex' in hello[5]:
            ws['C7'] = "Vac"
            ws['D7'] = "Vac"
        if 'Alex' in hello[6]:
            ws['C8'] = "Vac"
            ws['D8'] = "Vac"
        if 'Alex' in hello[7]:
            ws['C9'] = "Vac"
            ws['D9'] = "Vac"
        if 'Alex' in hello[8]:
            ws['C10'] = "Vac"
            ws['D10'] = "Vac"
        if 'Alex' in hello[9]:
            ws['C11'] = "Vac"
            ws['D11'] = "Vac"
    if 'AlexVacTue' in val:
        if 'Alex' in hello[0]:
            ws['E2'] = "Vac"
            ws['F2'] = "Vac"
        if 'Alex' in hello[1]:
            ws['E3'] = "Vac"
            ws['F3'] = "Vac"
        if 'Alex' in hello[2]:
            ws['E4'] = "Vac"
            ws['F4'] = "Vac"
        if 'Alex' in hello[3]:
            ws['E5'] = "Vac"
            ws['F5'] = "Vac"
        if 'Alex' in hello[4]:
            ws['E6'] = "Vac"
            ws['F6'] = "Vac"
        if 'Alex' in hello[5]:
            ws['E7'] = "Vac"
            ws['F7'] = "Vac"
        if 'Alex' in hello[6]:
            ws['E8'] = "Vac"
            ws['F8'] = "Vac"
        if 'Alex' in hello[7]:
            ws['E9'] = "Vac"
            ws['F9'] = "Vac"
        if 'Alex' in hello[8]:
            ws['E10'] = "Vac"
            ws['F10'] = "Vac"
        if 'Alex' in hello[9]:
            ws['E11'] = "Vac"
            ws['F11'] = "Vac"
    if 'AlexVacWed' in val:
        if 'Alex' in hello[0]:
            ws['G2'] = "Vac"
            ws['H2'] = "Vac"
        if 'Alex' in hello[1]:
            ws['G3'] = "Vac"
            ws['H3'] = "Vac"
        if 'Alex' in hello[2]:
            ws['G4'] = "Vac"
            ws['H4'] = "Vac"
        if 'Alex' in hello[3]:
            ws['G5'] = "Vac"
            ws['H5'] = "Vac"
        if 'Alex' in hello[4]:
            ws['G6'] = "Vac"
            ws['H6'] = "Vac"
        if 'Alex' in hello[5]:
            ws['G7'] = "Vac"
            ws['H7'] = "Vac"
        if 'Alex' in hello[6]:
            ws['G8'] = "Vac"
            ws['H8'] = "Vac"
        if 'Alex' in hello[7]:
            ws['G9'] = "Vac"
            ws['H9'] = "Vac"
        if 'Alex' in hello[8]:
            ws['G10'] = "Vac"
            ws['H10'] = "Vac"
        if 'Alex' in hello[9]:
            ws['G11'] = "Vac"
            ws['H11'] = "Vac"
    if 'AlexVacThu' in val:
        if 'Alex' in hello[0]:
            ws['I2'] = "Vac"
            ws['j2'] = "Vac"
        if 'Alex' in hello[1]:
            ws['I3'] = "Vac"
            ws['j3'] = "Vac"
        if 'Alex' in hello[2]:
            ws['I4'] = "Vac"
            ws['j4'] = "Vac"
        if 'Alex' in hello[3]:
            ws['I5'] = "Vac"
            ws['j5'] = "Vac"
        if 'Alex' in hello[4]:
            ws['I6'] = "Vac"
            ws['j6'] = "Vac"
        if 'Alex' in hello[5]:
            ws['I7'] = "Vac"
            ws['j7'] = "Vac"
        if 'Alex' in hello[6]:
            ws['I8'] = "Vac"
            ws['j8'] = "Vac"
        if 'Alex' in hello[7]:
            ws['I9'] = "Vac"
            ws['j9'] = "Vac"
        if 'Alex' in hello[8]:
            ws['I10'] = "Vac"
            ws['j10'] = "Vac"
        if 'Alex' in hello[9]:
            ws['I11'] = "Vac"
            ws['j11'] = "Vac"
    if 'AlexVacFri' in val:
        if 'Alex' in hello[0]:
            ws['K2'] = "Vac"
        if 'Alex' in hello[1]:
            ws['K3'] = "Vac"
        if 'Alex' in hello[2]:
            ws['K4'] = "Vac"
        if 'Alex' in hello[3]:
            ws['K5'] = "Vac"
        if 'Alex' in hello[4]:
            ws['K6'] = "Vac"
        if 'Alex' in hello[5]:
            ws['K7'] = "Vac"
        if 'Alex' in hello[6]:
            ws['K8'] = "Vac"
        if 'Alex' in hello[7]:
            ws['K9'] = "Vac"
        if 'Alex' in hello[8]:
            ws['K10'] = "Vac"
        if 'Alex' in hello[9]:
            ws['K11'] = "Vac"
    if 'AlexVacSat' in val:
        if 'Alex' in hello[0]:
            ws['L2'] = "Vac"
        if 'Alex' in hello[1]:
            ws['L3'] = "Vac"
        if 'Alex' in hello[2]:
            ws['L4'] = "Vac"
        if 'Alex' in hello[3]:
            ws['L5'] = "Vac"
        if 'Alex' in hello[4]:
            ws['L6'] = "Vac"
        if 'Alex' in hello[5]:
            ws['L7'] = "Vac"
        if 'Alex' in hello[6]:
            ws['L8'] = "Vac"
        if 'Alex' in hello[7]:
            ws['L9'] = "Vac"
        if 'Alex' in hello[8]:
            ws['L10'] = "Vac"
        if 'Alex' in hello[9]:
            ws['L11'] = "Vac"
    if 'AlexVacAll' in val:
        if 'Alex' in hello[0]:
            ws['A2'] = "Alex-Vac"
            ws['B2'] = "Alex-Vac"
            ws['C2'] = "Alex-Vac"
            ws['D2'] = "Alex-Vac"
            ws['E2'] = "Alex-Vac"
            ws['F2'] = "Alex-Vac"
            ws['G2'] = "Alex-Vac"
            ws['H2'] = "Alex-Vac"
            ws['I2'] = "Alex-Vac"
            ws['J2'] = "Alex-Vac"
            ws['K2'] = "Alex-Vac"
            ws['L2'] = "Alex-Vac"
        if 'Alex' in hello[1]:
            ws['A3'] = "Alex-Vac"
            ws['B3'] = "Alex-Vac"
            ws['C3'] = "Alex-Vac"
            ws['D3'] = "Alex-Vac"
            ws['E3'] = "Alex-Vac"
            ws['F3'] = "Alex-Vac"
            ws['G3'] = "Alex-Vac"
            ws['H3'] = "Alex-Vac"
            ws['I3'] = "Alex-Vac"
            ws['J3'] = "Alex-Vac"
            ws['K3'] = "Alex-Vac"
            ws['L3'] = "Alex-Vac"
        if 'Alex' in hello[2]:
            ws['A4'] = "Alex-Vac"
            ws['B4'] = "Alex-Vac"
            ws['C4'] = "Alex-Vac"
            ws['D4'] = "Alex-Vac"
            ws['E4'] = "Alex-Vac"
            ws['F4'] = "Alex-Vac"
            ws['G4'] = "Alex-Vac"
            ws['H4'] = "Alex-Vac"
            ws['I4'] = "Alex-Vac"
            ws['J4'] = "Alex-Vac"
            ws['K4'] = "Alex-Vac"
            ws['L4'] = "Alex-Vac"
        if 'Alex' in hello[3]:
            ws['A5'] = "Alex-Vac"
            ws['B5'] = "Alex-Vac"
            ws['C5'] = "Alex-Vac"
            ws['D5'] = "Alex-Vac"
            ws['E5'] = "Alex-Vac"
            ws['F5'] = "Alex-Vac"
            ws['G5'] = "Alex-Vac"
            ws['H5'] = "Alex-Vac"
            ws['I5'] = "Alex-Vac"
            ws['J5'] = "Alex-Vac"
            ws['K5'] = "Alex-Vac"
            ws['L5'] = "Alex-Vac"
        if 'Alex' in hello[4]:
            ws['A6'] = "Alex-Vac"
            ws['B6'] = "Alex-Vac"
            ws['C6'] = "Alex-Vac"
            ws['D6'] = "Alex-Vac"
            ws['E6'] = "Alex-Vac"
            ws['F6'] = "Alex-Vac"
            ws['G6'] = "Alex-Vac"
            ws['H6'] = "Alex-Vac"
            ws['I6'] = "Alex-Vac"
            ws['J6'] = "Alex-Vac"
            ws['K6'] = "Alex-Vac"
            ws['L6'] = "Alex-Vac"
        if 'Alex' in hello[5]:
            ws['A7'] = "Alex-Vac"
            ws['B7'] = "Alex-Vac"
            ws['C7'] = "Alex-Vac"
            ws['D7'] = "Alex-Vac"
            ws['E7'] = "Alex-Vac"
            ws['F7'] = "Alex-Vac"
            ws['G7'] = "Alex-Vac"
            ws['H7'] = "Alex-Vac"
            ws['I7'] = "Alex-Vac"
            ws['J7'] = "Alex-Vac"
            ws['K7'] = "Alex-Vac"
            ws['L7'] = "Alex-Vac"
        if 'Alex' in hello[6]:
            ws['A8'] = "Alex-Vac"
            ws['B8'] = "Alex-Vac"
            ws['C8'] = "Alex-Vac"
            ws['D8'] = "Alex-Vac"
            ws['E8'] = "Alex-Vac"
            ws['F8'] = "Alex-Vac"
            ws['G8'] = "Alex-Vac"
            ws['H8'] = "Alex-Vac"
            ws['I8'] = "Alex-Vac"
            ws['J8'] = "Alex-Vac"
            ws['K8'] = "Alex-Vac"
            ws['L8'] = "Alex-Vac"
        if 'Alex' in hello[7]:
            ws['A9'] = "Alex-Vac"
            ws['B9'] = "Alex-Vac"
            ws['C9'] = "Alex-Vac"
            ws['D9'] = "Alex-Vac"
            ws['E9'] = "Alex-Vac"
            ws['F9'] = "Alex-Vac"
            ws['G9'] = "Alex-Vac"
            ws['H9'] = "Alex-Vac"
            ws['I9'] = "Alex-Vac"
            ws['J9'] = "Alex-Vac"
            ws['K9'] = "Alex-Vac"
            ws['L9'] = "Alex-Vac"
        if 'Alex' in hello[8]:
            ws['A9'] = "Alex-Vac"
            ws['B9'] = "Alex-Vac"
            ws['C9'] = "Alex-Vac"
            ws['D9'] = "Alex-Vac"
            ws['E9'] = "Alex-Vac"
            ws['F9'] = "Alex-Vac"
            ws['G9'] = "Alex-Vac"
            ws['H9'] = "Alex-Vac"
            ws['I9'] = "Alex-Vac"
            ws['J9'] = "Alex-Vac"
            ws['K9'] = "Alex-Vac"
            ws['L9'] = "Alex-Vac"
        if 'Alex' in hello[9]:
            ws['A10'] = "Alex-Vac"
            ws['B10'] = "Alex-Vac"
            ws['C10'] = "Alex-Vac"
            ws['D10'] = "Alex-Vac"
            ws['E10'] = "Alex-Vac"
            ws['F10'] = "Alex-Vac"
            ws['G10'] = "Alex-Vac"
            ws['H10'] = "Alex-Vac"
            ws['I10'] = "Alex-Vac"
            ws['J10'] = "Alex-Vac"
            ws['K10'] = "Alex-Vac"
            ws['L10'] = "Alex-Vac"
    # excel design
    ws['A1'].style = highlight
    ws['A2'].style = highlight
    ws['A3'].style = highlight
    ws['A4'].style = highlight
    ws['A5'].style = highlight
    ws['A6'].style = highlight
    ws['A7'].style = highlight
    ws['A8'].style = highlight
    ws['A9'].style = highlight
    ws['A10'].style = highlight
    ws['A11'].style = highlight
    ws['B1'].style = highlight
    ws['B2'].style = highlight
    ws['B3'].style = highlight
    ws['B4'].style = highlight
    ws['B5'].style = highlight
    ws['B6'].style = highlight
    ws['B7'].style = highlight
    ws['B8'].style = highlight
    ws['B9'].style = highlight
    ws['B10'].style = highlight
    ws['B11'].style = highlight
    ws['C1'].style = highlight
    ws['C2'].style = highlight
    ws['C3'].style = highlight
    ws['C4'].style = highlight
    ws['C5'].style = highlight
    ws['C6'].style = highlight
    ws['C7'].style = highlight
    ws['C8'].style = highlight
    ws['C9'].style = highlight
    ws['C10'].style = highlight
    ws['C11'].style = highlight
    ws['D1'].style = highlight
    ws['D2'].style = highlight
    ws['D3'].style = highlight
    ws['D4'].style = highlight
    ws['D5'].style = highlight
    ws['D6'].style = highlight
    ws['D7'].style = highlight
    ws['D8'].style = highlight
    ws['D9'].style = highlight
    ws['D10'].style = highlight
    ws['D11'].style = highlight
    ws['E1'].style = highlight
    ws['E2'].style = highlight
    ws['E3'].style = highlight
    ws['E4'].style = highlight
    ws['E5'].style = highlight
    ws['E6'].style = highlight
    ws['E7'].style = highlight
    ws['E8'].style = highlight
    ws['E9'].style = highlight
    ws['E10'].style = highlight
    ws['E11'].style = highlight
    ws['F1'].style = highlight
    ws['F2'].style = highlight
    ws['F3'].style = highlight
    ws['F4'].style = highlight
    ws['F5'].style = highlight
    ws['F6'].style = highlight
    ws['F7'].style = highlight
    ws['F8'].style = highlight
    ws['F9'].style = highlight
    ws['F10'].style = highlight
    ws['F11'].style = highlight
    ws['G1'].style = highlight
    ws['G2'].style = highlight
    ws['G3'].style = highlight
    ws['G4'].style = highlight
    ws['G5'].style = highlight
    ws['G6'].style = highlight
    ws['G7'].style = highlight
    ws['G8'].style = highlight
    ws['G9'].style = highlight
    ws['G10'].style = highlight
    ws['G11'].style = highlight
    ws['H1'].style = highlight
    ws['H2'].style = highlight
    ws['H3'].style = highlight
    ws['H4'].style = highlight
    ws['H5'].style = highlight
    ws['H6'].style = highlight
    ws['H7'].style = highlight
    ws['H8'].style = highlight
    ws['H9'].style = highlight
    ws['H10'].style = highlight
    ws['H11'].style = highlight
    ws['I1'].style = highlight
    ws['I2'].style = highlight
    ws['I3'].style = highlight
    ws['I4'].style = highlight
    ws['I5'].style = highlight
    ws['I6'].style = highlight
    ws['I7'].style = highlight
    ws['I8'].style = highlight
    ws['I9'].style = highlight
    ws['I10'].style = highlight
    ws['I11'].style = highlight
    ws['J1'].style = highlight
    ws['J2'].style = highlight
    ws['J3'].style = highlight
    ws['J4'].style = highlight
    ws['J5'].style = highlight
    ws['J6'].style = highlight
    ws['J7'].style = highlight
    ws['J8'].style = highlight
    ws['J9'].style = highlight
    ws['J10'].style = highlight
    ws['J11'].style = highlight
    ws['K1'].style = highlight
    ws['K2'].style = highlight
    ws['K3'].style = highlight
    ws['K4'].style = highlight
    ws['K5'].style = highlight
    ws['K6'].style = highlight
    ws['K7'].style = highlight
    ws['K8'].style = highlight
    ws['K9'].style = highlight
    ws['K10'].style = highlight
    ws['K11'].style = highlight
    ws['L1'].style = highlight
    ws['L2'].style = highlight
    ws['L3'].style = highlight
    ws['L4'].style = highlight
    ws['L5'].style = highlight
    ws['L6'].style = highlight
    ws['L7'].style = highlight
    ws['L8'].style = highlight
    ws['L9'].style = highlight
    ws['L10'].style = highlight
    ws['L11'].style = highlight
    ws['A1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['B1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['C1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['D1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['E1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['F1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['G1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['H1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['I1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['J1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['K1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['L1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['B2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    wb.save('shifts.xlsx')

def button_clear():
    global val
    global Alyona
    global Alex
    global Ofir
    global Almog
    global Yair
    # global Pavel
    global Ran
    global Sahar
    global Alyona_st
    global Alex_st
    global Ofir_st
    global Yair_st
    global Almog_st
    global Ran_st
    # global Pavel_st
    global Sahar_st
    global Alyona_fr
    global Alex_fr
    global Ofir_fr
    global Yair_fr
    global Almog_fr
    global Ran_fr
    # global Pavel_fr
    global Sahar_fr
    global hello
    global maxst
    global maxfr
    global A
    global B
    global C
    global D
    global E
    global F
    global G
    global H
    global I
    global j

    A = ""
    B = ""
    C = ""
    D = ""
    E = ""
    F = ""
    G = ""
    H = ""
    I = ""
    J = ""
    val = ""
    Alyona = ""
    Alex = ""
    Ofir = ""
    Yair = ""
    Almog = ""
    # Pavel = ""
    Ran = ""
    Sahar = ""
    hello = ""
    maxst = ""
    maxfr = ""
    Alyona_st = 0
    Alyona_fr = 0
    Alex_st = 0
    Alex_fr = 0
    Ofir_st = 0
    Ofir_fr = 0
    Yair_st = 0
    Yair_fr = 0
    Almog_st = 0
    Almog_fr = 0
    Ran_st = 0
    Ran_fr = 0
    Sahar_st = 0
    Sahar_fr = 0
    # Pavel_st = 0
    # Pavel_fr = 0

    label_5 = customtkinter.CTkLabel(master=frame_1, text=("You have clear all request"))
    label_5.grid(pady=1, padx=1)

def his():
    ex = ["SaturDay", "FriDay"]
    data = [[hello[4], hello[5]], [hello[6], hello[7]], [hello[8], hello[9]]]
    df = pd.DataFrame(data, columns=ex)
    append_df_to_excel('history.xlsx', df, sheet_name='his', index=True, startrow=0)

def get_data():
    global Alyona_st
    global Alex_st
    global Ofir_st
    global Yair_st
    global Almog_st
    global Ran_st
    # global Pavel_st
    global Sahar_st
    global Alyona_fr
    global Alex_fr
    global Ofir_fr
    global Yair_fr
    global Almog_fr
    global Ran_fr
    # global Pavel_fr
    global Sahar_fr
    sheets_list = wb2.sheetnames
    # range (i, j) and if j < 30 delete first item
    for i in range(0, 100):
        if 0 <= i < len(sheets_list):
            df = pd.read_excel('history.xlsx', sheet_name=sheets_list[i])
            st = (df['SaturDay'])
            fr = (df['FriDay'])
            if 'Alyona' in st.values:
                Alyona_st = Alyona_st + 1
            if 'Alyona' in fr.values:
                Alyona_fr = Alyona_fr + 1
            if 'Alex' in st.values:
                Alex_st = Alex_st + 1
            if 'Alex' in fr.values:
                Alex_fr = Alex_fr + 1
            if 'Ofir' in st.values:
                Ofir_st = Ofir_st + 1
            if 'Ofir' in fr.values:
                Ofir_fr = Ofir_fr + 1
            if 'Yair' in st.values:
                Yair_st = Yair_st + 1
            if 'Yair' in fr.values:
                Yair_fr = Yair_fr + 1
            if 'Almog' in st.values:
                Almog_st = Almog_st + 1
            if 'Almog' in fr.values:
                Almog_fr = Almog_fr + 1
            if 'Ran' in st.values:
                Ran_st = Ran_st + 1
            if 'Ran' in fr.values:
                Ran_fr = Ran_fr + 1
            if 'Sahar' in st.values:
                Sahar_st = Sahar_st + 1
            if 'Sahar' in fr.values:
                Sahar_fr = Sahar_fr + 1
            # if 'Pavel' in st.values:
            #     Pavel_st = Pavel_st + 1
            # if 'Pavel' in fr.values:
            #     Pavel_fr = Pavel_fr + 1

def ai():
    print("aba")




def block():
    global temp
    global val
    temp = " "
    operator = optionmenu_1.get()
    if (combobox_1.get()) == 'Block':
        #alyona block
        if (switchvar1.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Sun-Morning '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Sun-Night '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Mon-Morning '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Mon-Night '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Tue-Morning '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Tue-Night '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Wed-Morning '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Wed-Night '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Thu-Morning '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Thu-Night '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Fri-Morning '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona Sat-Night '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Alyona':
            temp = temp + 'Alyona All'
            print(temp)
        #alex block
        if (switchvar1.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Sun-Morning '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Sun-Night '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Mon-Morning '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Mon-Night '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Tue-Morning '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Tue-Night '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Wed-Morning '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Wed-Night '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Thu-Morning '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Thu-Night '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Fri-Morning '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex Sat-Night '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Alex':
            temp = temp + 'Alex All'
            print(temp)
        #ofir block
        if (switchvar1.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Sun-Morning '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Sun-Night '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Mon-Morning '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Mon-Night '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Tue-Morning '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Tue-Night '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Wed-Morning '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Wed-Night '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Thu-Morning '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Thu-Night '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Fri-Morning '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir Sat-Night '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Ofir':
            temp = temp + 'Ofir All'
            print(temp)
        #yair block
        if (switchvar1.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Sun-Morning '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Sun-Night '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Mon-Morning '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Mon-Night '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Tue-Morning '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Tue-Night '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Wed-Morning '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Wed-Night '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Thu-Morning '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Thu-Night '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Fri-Morning '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair Sat-Night '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Yair':
            temp = temp + 'Yair All'
            print(temp)
        #Almog block

        if (switchvar1.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Sun-Morning '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Sun-Night '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Mon-Morning '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Mon-Night '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Tue-Morning '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Tue-Night '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Wed-Morning '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Wed-Night '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Thu-Morning '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Thu-Night '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Fri-Morning '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog Sat-Night '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Almog':
            temp = temp + 'Almog All'
            print(temp)
        #ran block

        if (switchvar1.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Sun-Morning '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Sun-Night '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Mon-Morning '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Mon-Night '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Tue-Morning '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Tue-Night '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Wed-Morning '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Wed-Night '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Thu-Morning '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Thu-Night '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Fri-Morning '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran Sat-Night '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Ran':
            temp = temp + 'Ran All'
            print(temp)
        #sahar block
        if (switchvar1.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Sun-Morning '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Sun-Night '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Mon-Morning '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Mon-Night '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Tue-Morning '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Tue-Night '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Wed-Morning '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Wed-Night '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Thu-Morning '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Thu-Night '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Fri-Morning '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar Sat-Night '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Sahar':
            temp = temp + 'Sahar All'
            print(temp)
    if (combobox_1.get()) == 'Vac':
        # alyona vac
        if (switchvar1.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacSun '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacSun '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacMon '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacMon '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacTue '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacTue '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacWed '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacWed '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacThu '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacThu '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacFri '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVacSat '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Alyona':
            temp = temp + 'AlyonaVac All'
            print(temp)
        # alex vac
        if (switchvar1.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacSun '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacSun '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacMon '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacMon '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacTue '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacTue '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacWed '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacWed '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacThu '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacThu '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacFri '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVacSat '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Alex':
            temp = temp + 'AlexVac All'
            print(temp)
        # ofir vac
        if (switchvar1.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacSun '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacSun '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacMon '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacMon '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacTue '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacTue '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacWed '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacWed '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacThu '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacThu '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacFri '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacSat '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Ofir':
            temp = temp + 'OfirVacAll'
            print(temp)
        # yair vac
        if (switchvar1.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacSun '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacSun '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacMon '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacMon '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacTue '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacTue '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacWed '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacWed '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacThu '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacThu '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacFri '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacSat '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Yair':
            temp = temp + 'YairVacAll'
            print(temp)
        # Almog vac
        if (switchvar1.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacSun '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacSun'
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacMon '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacMon '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacTue '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacTue '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacWed '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacWed '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacThu '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacThu '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacFri '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacSat '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Almog':
            temp = temp + 'AlmogVacAll '
            print(temp)
        # Ran vac
        if (switchvar1.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacSun '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacSun '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacMon '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacMon '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacTue '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacTue '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacWed '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacWed '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacThu '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacThu '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacFri '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacSat '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Ran':
            temp = temp + 'RanVacAll'
            print(temp)
        # Sahar vac
        if (switchvar1.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacSun '
            print(temp)
        if (switchvar2.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacSun '
            print(temp)
        if (switchvar3.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacMon '
            print(temp)
        if (switchvar4.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacMon '
            print(temp)
        if (switchvar5.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacTue '
            print(temp)
        if (switchvar6.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacTue '
            print(temp)
        if (switchvar7.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacWed '
            print(temp)
        if (switchvar8.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacWed '
            print(temp)
        if (switchvar9.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacThu '
            print(temp)
        if (switchvar10.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacThu '
            print(temp)
        if (switchvar11.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacFri '
            print(temp)
        if (switchvar12.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacSat '
            print(temp)
        if (selectvar.get()) == 1 and operator == 'Sahar':
            temp = temp + 'SaharVacAll'
            print(temp)


def button_save():
    global temp
    global val
    val = val + temp
    print(val)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='new')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


frame_1 = customtkinter.CTkFrame(master=app)
frame_1.grid(pady=60, padx=60)

label_1 = customtkinter.CTkLabel(master=frame_1, text="Shifts Creator", justify=tkinter.LEFT)
label_1.grid(row = 0, column = 0, pady=12, padx=10)



optionmenu_1 = customtkinter.CTkOptionMenu(frame_1,
                                           values=["Alyona", "Alex", "Ofir", "Yair", "Almog", "Ran", "Sahar"])
optionmenu_1.grid(row = 1, column = 0, pady=12, padx=10)
optionmenu_1.set("Alyona")

# radio_var = tkinter.IntVar()
#
# radio_button_1 = customtkinter.CTkRadioButton(master=frame_1,
#
#
#                                                           variable=radio_var)
combobox_1 = customtkinter.CTkComboBox(frame_1, values=["Block","Vac"])
combobox_1.grid(sticky = W,row = 2, column = 0,pady=12, padx=10)

selectvar = IntVar()


switch_13 = customtkinter.CTkSwitch(master=frame_1,text="Select All", command=block, variable=selectvar,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 2,column = 1,pady=9, padx=8)

switchvar1 = IntVar()
switchvar2 = IntVar()
switchvar3 = IntVar()
switchvar4 = IntVar()
switchvar5 = IntVar()
switchvar6 = IntVar()
switchvar7 = IntVar()
switchvar8 = IntVar()
switchvar9 = IntVar()
switchvar10 = IntVar()
switchvar11 = IntVar()
switchvar12 = IntVar()



switch_1 = customtkinter.CTkSwitch(master=frame_1,text="Sun-Morning", command=block, variable=switchvar1,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 3, column = 0, pady=8, padx=8)

switch_2 = customtkinter.CTkSwitch(master=frame_1,text="Sun-Night", command=block, variable=switchvar2,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 3, column = 1,pady=8, padx=8)

switch_3 = customtkinter.CTkSwitch(master=frame_1,text="Mon-Morning", command=block, variable=switchvar3,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 4, column = 0,pady=8, padx=8)

switch_3 = customtkinter.CTkSwitch(master=frame_1,text="Mon-Night", command=block, variable=switchvar4,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 4, column = 1,pady=8, padx=8)

switch_4 = customtkinter.CTkSwitch(master=frame_1,text="Tue-Morning", command=block, variable=switchvar5,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 5, column = 0,pady=8, padx=8)

switch_8 = customtkinter.CTkSwitch(master=frame_1,text="Tue-Night", command=block, variable=switchvar6,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 5, column = 1,pady=8, padx=8)

switch_7 = customtkinter.CTkSwitch(master=frame_1,text="Wed-Morning", command=block, variable=switchvar7,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 6, column = 0,pady=8, padx=8)

switch_8 = customtkinter.CTkSwitch(master=frame_1,text="Wed-Night", command=block, variable=switchvar8,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 6, column = 1,pady=8, padx=8)

switch_9 = customtkinter.CTkSwitch(master=frame_1,text="Thu-Morning", command=block, variable=switchvar9,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 7, column = 0,pady=8, padx=8)

switch_10 = customtkinter.CTkSwitch(master=frame_1,text="Thu-Night", command=block, variable=switchvar10,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 7, column = 1,pady=8, padx=8)

switch_11 = customtkinter.CTkSwitch(master=frame_1,text="Fri-Morning", command=block, variable=switchvar11,
                                                 onvalue=1, offvalue=0).grid(sticky = W,row = 8, column = 0,pady=8, padx=8)

switch_12 = customtkinter.CTkSwitch(master=frame_1,text="Sat-Night", command=block, variable=switchvar12,
                                                 onvalue=1, offvalue=0).grid(sticky = W, row = 8, column = 1,pady=8, padx=8)






# radio_button_1.grid(pady=12, padx=10)
#
# radio_button_1.select()

# v = StringVar(frame_1, "1")
#
# style = Style(frame_1)
# style.configure("TRadiobutton", background="gray",
#                 foreground="light blue", font=("arial", 10, "bold"))
#
# Radiobutton(frame_1, text='text', variable=v,
#             value=0).grid(side=TOP, ipady=5)


# combobox_1.set("Sun-Morning")

button_1 = customtkinter.CTkButton(master=frame_1, text="save", command=button_save)
button_1.grid(pady=12, padx=10)

button_2 = customtkinter.CTkButton(master=frame_1, text="finished", command=button_callback2)
button_2.grid(pady=12, padx=10)

# button_3 = customtkinter.CTkButton(master=frame_1, text="Clear all", command=button_clear)
# button_3.grid(pady=10, padx=8)

label_3 = customtkinter.CTkLabel(master=frame_1, text="Shifts blocked:", justify=tkinter.LEFT)
label_3.grid(pady=5, padx=3)






app.mainloop()
