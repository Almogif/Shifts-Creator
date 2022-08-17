import tkinter
import customtkinter
import random
from openpyxl import Workbook



customtkinter.set_appearance_mode("dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

app = customtkinter.CTk()
app.geometry("400x580")
app.title("Shifts creator")


wb = Workbook()
ws = wb.active
ws.append(['Sun-Day', 'Sun-Night','Mon-Day','Mon-Night','Tue-Day','Tue-Night','Wed-Day','Wed-Night','Thu-Day','Thu-Night','Fri-Day','Sat-Night'])




val = ""
alyona = ""
alex = ""
ofir = ""
yair = ""
almog = ""
pavel = ""
ran = ""
sahar = ""
hello = ""

shifts = "shifts: "
i = 0
count = 0

def button_callback():
    global val
    val = (" " + val + (optionmenu_1.get()) + " " + (combobox_1.get()) + " ")
    print(val)
    label_2 = customtkinter.CTkLabel(master=frame_1,
                                     text=" " + (optionmenu_1.get()) + " can not work on " + (combobox_1.get()) + " ")
    label_2.pack(pady=1, padx=1)


def button_callback2():
    global val
    global alyona
    global alex
    global ofir
    global almog
    global yair
    global pavel
    global ran
    global sahar
    global RandomA
    global RandomB
    global RandomC
    global RandomD
    global RandomE
    global RandomF
    global RandomG
    global RandomH
    global RandomI
    global RandomJ
    global shifts
    # Alyona check
    if "Alyona Sun-Morning" in val:
        alyona = alyona + "ABCD"
    if "Alyona Mon-Morning" in val:
        alyona = alyona + "ABCDG"
    if "Alyona Tue-Morning" in val:
        alyona = alyona + "ABGIJ"
    if "Alyona Wed-Morning" in val:
        alyona = alyona + "BEFIJ"
    if "Alyona Thu-Morning" in val:
        alyona = alyona + "EFHJ"
    if "Alyona Fri-Morning " in val:
        alyona = alyona + "EFH"
    if "Alyona Sun-Night" in val:
        alyona = alyona + "EFHIJ"
    if "Alyona Mon-Night" in val:
        alyona = alyona + "EFH"
    if "Alyona Tue-Night" in val:
        alyona = alyona + "CDH"
    if "Alyona Wed-Night" in val:
        alyona = alyona + "ACDG"
    if "Alyona Thu-Night" in val:
        alyona = alyona + "ABDGI"
    if "Alyona Sat-Night " in val:
        alyona = alyona + "CGIJ"
    # alex cheack
    if "Alex Sun-Morning" in val:
        alex = alex + "ABCD"
        print("alex check")
    if "Alex Mon-Morning" in val:
        alex = alex + "ABCDG"
    if "Alex Tue-Morning" in val:
        alex = alex + "ABGIJ"
    if "Alex Wed-Morning" in val:
        alex = alex + "BEFIJ"
    if "Alex Thu-Morning" in val:
        alex = alex + "EFHJ"
    if "Alex Fri-Morning " in val:
        alex = alex + "EFH"
    if "Alex Sun-Night" in val:
        alex = alex + "EFHIJ"
    if "Alex Mon-Night" in val:
        alex = alex + "EFH"
    if "Alex Tue-Night" in val:
        alex = alex + "CDH"
    if "Alex Wed-Night" in val:
        alex = alex + "ACDG"
    if "Alex Thu-Night" in val:
        alex = alex + "ABDGI"
    if "Alex Sat-Night " in val:
        alex = alex + "CGIJ"
    # ofir check
    if "Ofir Sun-Morning" in val:
        ofir = ofir + "ABCD"
    if "Ofir Mon-Morning" in val:
        ofir = ofir + "ABCDG"
    if "Ofir Tue-Morning" in val:
        ofir = ofir + "ABGIJ"
    if "Ofir Wed-Morning" in val:
        ofir = ofir + "BEFIJ"
    if "Ofir Thu-Morning" in val:
        ofir = ofir + "EFHJ"
    if "Ofir Fri-Morning " in val:
        ofir = ofir + "EFH"
    if "Ofir Sun-Night" in val:
        ofir = ofir + "EFHIJ"
    if "Ofir Mon-Night" in val:
        ofir = ofir + "EFH"
    if "Ofir Tue-Night" in val:
        ofir = ofir + "CDH"
    if "Ofir Wed-Night" in val:
        ofir = ofir + "ACDG"
    if "Ofir Thu-Night" in val:
        ofir = ofir + "ABDGI"
    if "Ofir Sat-Night " in val:
        ofir = ofir + "CGIJ"
    # yair check
    if "Yair Sun-Morning" in val:
        yair = yair + "ABCD"
    if "Yair Mon-Morning" in val:
        yair = yair + "ABCDG"
    if "Yair Tue-Morning" in val:
        yair = yair + "ABGIJ"
    if "Yair Wed-Morning" in val:
        yair = yair + "BEFIJ"
    if "Yair Thu-Morning" in val:
        yair = yair + "EFHJ"
    if "Yair Fri-Morning " in val:
        yair = yair + "EFH"
    if "Yair Sun-Night" in val:
        yair = yair + "EFHIJ"
    if "Yair Mon-Night" in val:
        yair = yair + "EFH"
    if "Yair Tue-Night" in val:
        yair = yair + "CDH"
    if "Yair Wed-Night" in val:
        yair = yair + "ACDG"
    if "Yair Thu-Night" in val:
        yair = yair + "ABDGI"
    if "Yair Sat-Night " in val:
        yair = yair + "CGIJ"
    # pavel check
    if "Pavel Sun-Morning" in val:
        pavel = pavel + "ABCD"
    if "Pavel Mon-Morning" in val:
        pavel = pavel + "ABCDG"
    if "Pavel Tue-Morning" in val:
        pavel = pavel + "ABGIJ"
    if "Pavel Wed-Morning" in val:
        pavel = pavel + "BEFIJ"
    if "Pavel Thu-Morning" in val:
        pavel = pavel + "EFHJ"
    if "Pavel Fri-Morning " in val:
        pavel = pavel + "EFH"
    if "Pavel Sun-Night" in val:
        pavel = pavel + "EFHIJ"
    if "Pavel Mon-Night" in val:
        pavel = pavel + "EFH"
    if "Pavel Tue-Night" in val:
        pavel = pavel + "CDH"
    if "Pavel Wed-Night" in val:
        pavel = pavel + "ACDG"
    if "Pavel Thu-Night" in val:
        pavel = pavel + "ABDGI"
    if "Pavel Sat-Night " in val:
        pavel = pavel + "CGIJ"
        # almog check
    if "Almog Sun-Morning" in val:
        almog = almog + "ABCD"
    if "Almog Mon-Morning" in val:
        almog = almog + "ABCDG"
    if "Almog Tue-Morning" in val:
        almog = almog + "ABGIJ"
    if "Almog Wed-Morning" in val:
        almog = almog + "BEFIJ"
    if "Almog Thu-Morning" in val:
        almog = almog + "EFHJ"
    if "Almog Fri-Morning " in val:
        almog = almog + "EFH"
    if "Almog Sun-Night" in val:
        almog = almog + "EFHIJ"
    if "Almog Mon-Night" in val:
        almog = almog + "EFH"
    if "Almog Tue-Night" in val:
        almog = almog + "CDH"
    if "Almog Wed-Night" in val:
        almog = almog + "ACDG"
    if "Almog Thu-Night" in val:
        almog = almog + "ABDGI"
    if "Almog Sat-Night " in val:
        almog = almog + "CGIJ"
        # ran check
    if "Ran Sun-Morning" in val:
        ran = ran + "ABCD"
    if "Ran Mon-Morning" in val:
        ran = ran + "ABCDG"
    if "Ran Tue-Morning" in val:
        ran = ran + "ABGIJ"
    if "Ran Wed-Morning" in val:
        ran = ran + "BEFIJ"
    if "Ran Thu-Morning" in val:
        ran = ran + "EFHJ"
    if "Ran Fri-Morning " in val:
        ran = ran + "EFH"
    if "Ran Sun-Night" in val:
        ran = ran + "EFHIJ"
    if "Ran Mon-Night" in val:
        ran = ran + "EFH"
    if "Ran Tue-Night" in val:
        ran = ran + "CDH"
    if "Ran Wed-Night" in val:
        ran = ran + "ACDG"
    if "Ran Thu-Night" in val:
        ran = ran + "ABDGI"
    if "Ran Sat-Night " in val:
        ran = ran + "CGIJ"
        # sahar check
    if "Sahar Sun-Morning" in val:
        sahar = sahar + "ABCD"
    if "Sahar Mon-Morning" in val:
        sahar = sahar + "ABCDG"
    if "Sahar Tue-Morning" in val:
        sahar = sahar + "ABGIJ"
    if "Sahar Wed-Morning" in val:
        sahar = sahar + "BEFIJ"
    if "Sahar Thu-Morning" in val:
        sahar = sahar + "EFHJ"
    if "Sahar Fri-Morning " in val:
        sahar = sahar + "EFH"
    if "Sahar Sun-Night" in val:
        sahar = sahar + "EFHIJ"
    if "Sahar Mon-Night" in val:
        sahar = sahar + "EFH"
    if "Sahar Tue-Night" in val:
        sahar = sahar + "CDH"
    if "Sahar Wed-Night" in val:
        sahar = sahar + "ACDG"
    if "Sahar Thu-Night" in val:
        sahar = sahar + "ABDGI"
    if "Sahar Sat-Night " in val:
        sahar = sahar + "CGIJ"

    #shfits lists
    A = ""
    B = ""
    C = ""
    D = ""
    E = ""
    F = ""
    G = ""
    H = ""
    I = ""
    J = ""

    for x in range(0, 8):
        if x == 0:
            if not "A" in alyona:
                A = A + "alyona "
        elif x == 1:
            if not "A" in alex:
                A = A + "alex "
        elif x == 2:
            if not "A" in ofir:
                A = A + "ofir "
        elif x == 3:
            if not "A" in yair:
                A = A + "yair "
        elif x == 4:
            if not "A" in almog:
                A = A + "almog "
        elif x == 5:
            if not "A" in pavel:
                A = A + "pavel "
        elif x == 6:
            if not "A" in sahar:
                A = A + "sahar "
        elif x == 7:
            if not "A" in ran:
                A = A + "ran "
        if x == 0:
            if not "B" in alyona:
                B = B + "alyona "
        elif x == 1:
            if not "B" in alex:
                B = B + "alex "
        elif x == 2:
            if not "B" in ofir:
                B = B + "ofir "
        elif x == 3:
            if not "B" in yair:
                B = B + "yair "
        elif x == 4:
            if not "B" in almog:
                B = B + "almog "
        elif x == 5:
            if not "B" in pavel:
                B = B + "pavel "
        elif x == 6:
            if not "B" in sahar:
                B = B + "sahar "
        elif x == 7:
            if not "B" in ran:
                B = B + "ran "
        if x == 0:
            if not "C" in alyona:
                C = C + "alyona "
        elif x == 1:
            if not "C" in alex:
                C = C + "alex "
        elif x == 2:
            if not "C" in ofir:
                C = C + "ofir "
        elif x == 3:
            if not "C" in yair:
                C = C + "yair "
        elif x == 4:
            if not "C" in almog:
                C = C + "almog "
        elif x == 5:
            if not "C" in pavel:
                C = C + "pavel "
        elif x == 6:
            if not "C" in sahar:
                C = C + "sahar "
        elif x == 7:
            if not "C" in ran:
                C = C + "ran "
        if x == 0:
            if not "D" in alyona:
                D = D + "alyona "
        elif x == 1:
            if not "D" in alex:
                D = D + "alex "
        elif x == 2:
            if not "D" in ofir:
                D = D + "ofir "
        elif x == 3:
            if not "D" in yair:
                D = D + "yair "
        elif x == 4:
            if not "D" in almog:
                D = D + "almog "
        elif x == 5:
            if not "D" in pavel:
                D = D + "pavel "
        elif x == 6:
            if not "D" in sahar:
                D = D + "sahar "
        elif x == 7:
            if not "D" in ran:
                D = D + "ran "
        if x == 0:
            if not "E" in alyona:
                E = E + "alyona "
        elif x == 1:
            if not "E" in alex:
                E = E + "alex "
        elif x == 2:
            if not "E" in ofir:
                E = E + "ofir "
        elif x == 3:
            if not "E" in yair:
                E = E + "yair "
        elif x == 4:
            if not "E" in almog:
                E = E + "almog "
        elif x == 5:
            if not "E" in pavel:
                E = E + "pavel "
        elif x == 6:
            if not "E" in sahar:
                E = E + "sahar "
        elif x == 7:
            if not "E" in ran:
                E = E + "ran "
        if x == 0:
            if not "F" in alyona:
                F = F + "alyona "
        elif x == 1:
            if not "F" in alex:
                F = F + "alex "
        elif x == 2:
            if not "F" in ofir:
                F = F + "ofir "
        elif x == 3:
            if not "F" in yair:
                F = F + "yair "
        elif x == 4:
            if not "F" in almog:
                F = F + "almog "
        elif x == 5:
            if not "F" in pavel:
                F = F + "pavel "
        elif x == 6:
            if not "F" in sahar:
                F = F + "sahar "
        elif x == 7:
            if not "F" in ran:
                F = F + "ran "
        if x == 0:
            if not "G" in alyona:
                G = G + "alyona "
        elif x == 1:
            if not "G" in alex:
                G = G + "alex "
        elif x == 2:
            if not "G" in ofir:
                G = G + "ofir "
        elif x == 3:
            if not "G" in yair:
                G = G + "yair "
        elif x == 4:
            if not "G" in almog:
                G = G + "almog "
        elif x == 5:
            if not "G" in pavel:
                G = G + "pavel "
        elif x == 6:
            if not "G" in sahar:
                G = G + "sahar "
        elif x == 7:
            if not "G" in ran:
                G = G + "ran "
        if x == 0:
            if not "H" in alyona:
                H = H + "alyona "
        elif x == 1:
            if not "H" in alex:
                H = H + "alex "
        elif x == 2:
            if not "H" in ofir:
                H = H + "ofir "
        elif x == 3:
            if not "H" in yair:
                H = H + "yair "
        elif x == 4:
            if not "H" in almog:
                H = H + "almog "
        elif x == 5:
            if not "H" in pavel:
                H = H + "pavel "
        elif x == 6:
            if not "H" in sahar:
                H = H + "sahar "
        elif x == 7:
            if not "H" in ran:
                H = H + "ran "
        if x == 0:
            if not "I" in alyona:
                I = I + "alyona "
        elif x == 1:
            if not "I" in alex:
                I = I + "alex "
        elif x == 2:
            if not "I" in ofir:
                I = I + "ofir "
        elif x == 3:
            if not "I" in yair:
                I = I + "yair "
        elif x == 4:
            if not "I" in almog:
                I = I + "almog "
        elif x == 5:
            if not "I" in pavel:
                I = I + "pavel "
        elif x == 6:
            if not "I" in sahar:
                I = I + "sahar "
        elif x == 7:
            if not "I" in ran:
                I = I + "ran "
        if x == 0:
            if not "J" in alyona:
                J = J + "alyona "
        elif x == 1:
            if not "J" in alex:
                J = J + "alex "
        elif x == 2:
            if not "J" in ofir:
                J = J + "ofir "
        elif x == 3:
            if not "J" in yair:
                J = J + "yair "
        elif x == 4:
            if not "J" in almog:
                J = J + "almog "
        elif x == 5:
            if not "J" in pavel:
                J = J + "pavel "
        elif x == 6:
            if not "J" in sahar:
                J = J + "sahar "
        elif x == 7:
            if not "J" in ran:
                J = J + "ran "

        #Convert to list
        
    if A:
        A = A.split(" ")
        A.pop()
        print("wait")
    else:
        A = "empty empty empty"
        A = A.split(" ")
        print(A)
    if B:
        B = B.split(" ")
        B.pop()
    else:
        B = "empty empty empty"
        B = B.split(" ")
    if C:
        C = C.split(" ")
        C.pop()
    else:
        C = "empty empty empty"
        C = C.split(" ")
    if D:
        D = D.split(" ")
        D.pop()
    else:
        D = "empty empty empty"
        D = D.split(" ")
    if E:
        E = E.split(" ")
        E.pop()
    else:
        E = "empty empty empty"
        E = E.split(" ")
    if F:
        F = F.split(" ")
        F.pop()
    else:
        F = "empty empty empty"
        F = F.split(" ")
    if G:
        G = G.split(" ")
        G.pop()
    else:
        G = "empty empty empty"
        G = G.split(" ")
    if H:
        H = H.split(" ")
        H.pop()
    else:
        H = "empty empty empty"
        H = H.split(" ")
    if I:
        I = I.split(" ")
        I.pop()
    else:
        I = "empty empty empty"
        I = I.split(" ")
    if J:
        J = J.split(" ")
        J.pop()
    else:
        J = "empty empty empty"
        J = J.split(" ")

    global hello
    for x in range(0,9999):
        hello = random.choice(A) + " " + random.choice(B) + " " + random.choice(C) + " " + random.choice(
            D) + " " + random.choice(E) + " " + random.choice(F) + " " + random.choice(G) + " " + random.choice(
            H) + " " + random.choice(I) + " " + random.choice(J)

        hello = hello.split(" ")
        if (hello.count('empty') <= 2) and (hello.count('alyona') == 1) and (hello.count('alex') == 1) and (hello.count('ofir') == 1) and (hello.count('yair') == 1) and (hello.count('almog') == 2) and (hello.count('pavel') == 1) and (hello.count('ran') == 2) and (hello.count('sahar') == 1):
            break
        elif (hello.count('alyona') == 1) and (hello.count('alex') == 2) and (hello.count('ofir') == 1) and (hello.count('yair') == 2) and (hello.count('almog') == 1) and (hello.count('pavel') == 1) and (hello.count('ran') == 1) and (hello.count('sahar') == 1):
            break
        elif (hello.count('alyona') == 2) and (hello.count('alex') == 1) and (hello.count('ofir') == 1) and (hello.count('yair') == 1) and (hello.count('almog') == 1) and (hello.count('pavel') == 1) and (hello.count('ran') == 1) and (hello.count('sahar') == 2):
            break




    ws['A2'] = (hello[0])
    ws['A3'] = (hello[1])
    ws['A4'] = (hello[2])
    ws['A5'] = (hello[3])
    ws['C2'] = (hello[0])
    ws['C3'] = (hello[1])
    ws['C4'] = (hello[2])
    ws['C5'] = (hello[3])
    ws['E2'] = (hello[0])
    ws['E3'] = (hello[1])
    ws['F4'] = (hello[2])
    ws['F5'] = (hello[3])
    ws['H2'] = (hello[0])
    ws['G3'] = (hello[1])
    ws['H4'] = (hello[2])
    ws['H5'] = (hello[3])
    ws['J2'] = (hello[0])
    ws['L4'] = (hello[2])
    ws['J4'] = (hello[2])
    ws['J5'] = (hello[3])

    ws['B6'] = (hello[4])
    ws['B7'] = (hello[5])
    ws['C8'] = (hello[6])
    ws['B9'] = (hello[7])
    ws['B10'] = (hello[8])
    ws['B11'] = (hello[9])

    ws['D6'] = (hello[4])
    ws['D7'] = (hello[5])
    ws['E8'] = (hello[6])
    ws['D9'] = (hello[7])
    ws['E10'] = (hello[8])
    ws['E11'] = (hello[9])

    ws['G6'] = (hello[4])
    ws['G7'] = (hello[5])
    ws['H8'] = (hello[6])
    ws['F9'] = (hello[7])
    ws['G10'] = (hello[8])
    ws['G11'] = (hello[9])

    ws['I6'] = (hello[4])
    ws['I7'] = (hello[5])
    ws['J8'] = (hello[6])
    ws['I9'] = (hello[7])
    ws['J10'] = (hello[8])
    ws['I11'] = (hello[9])

    ws['K6'] = (hello[4])
    ws['K7'] = (hello[5])
    ws['L8'] = (hello[6])
    ws['K9'] = (hello[7])
    ws['L10'] = (hello[8])
    ws['L11'] = (hello[9])
    wb.save('sample.xlsx')




def button_clear():
    global val
    global shifts
    val = ""
    shifts = ""

    label_5 = customtkinter.CTkLabel(master=frame_1, text=("You have clear all request"))
    label_5.pack(pady=1, padx=1)


frame_1 = customtkinter.CTkFrame(master=app)
frame_1.pack(pady=60, padx=60, fill="both", expand=True)

label_1 = customtkinter.CTkLabel(master=frame_1, text="Shifts Creator", justify=tkinter.LEFT)
label_1.pack(pady=12, padx=10)

# set the text


# entry_1 = customtkinter.CTkEntry(master=frame_1, placeholder_text="CTkEntry")
# entry_1.pack(pady=12, padx=10)

optionmenu_1 = customtkinter.CTkOptionMenu(frame_1,
                                           values=["Alyona", "Alex", "Ofir", "Yair", "Almog", "Pavel", "Ran", "Sahar"])
optionmenu_1.pack(pady=12, padx=10)
optionmenu_1.set("Alyona")

combobox_1 = customtkinter.CTkComboBox(frame_1, values=["Sun-Morning", "Mon-Morning", "Tue-Morning", "Wed-Morning",
                                                        "Thu-Morning", "Fri-Morning", "Sun-Night", "Mon-Night",
                                                        "Tue-Night", "Wed-Night", "Thu-Night", "Sat-Night"])
combobox_1.pack(pady=12, padx=10)
combobox_1.set("Sun-Morning")

button_1 = customtkinter.CTkButton(master=frame_1, text="save", command=button_callback)
button_1.pack(pady=12, padx=10)

button_2 = customtkinter.CTkButton(master=frame_1, text="finished", command=button_callback2)
button_2.pack(pady=12, padx=10)

button_3 = customtkinter.CTkButton(master=frame_1, text="Clear all", command=button_clear)
button_3.pack(pady=10, padx=8)

label_3 = customtkinter.CTkLabel(master=frame_1, text="Shifts blocked:", justify=tkinter.LEFT)
label_3.pack(pady=5, padx=3)

app.mainloop()
