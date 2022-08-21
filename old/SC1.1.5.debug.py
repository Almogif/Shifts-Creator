import tkinter
import customtkinter
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl import load_workbook
import pandas as pd
import os
from pathlib import Path

highlight = NamedStyle(name="highlight")
bd = Side(style='thin', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = GradientFill(stop=("CEF0FA", "CEF0FA"))
highlight.font = Font(name='Tahoma',
                      size=11,
                      bold=True,
                      italic=False,
                      vertAlign=None,
                      underline='none',
                      strike=False,
                      color='FF000000')
fill = PatternFill(fill_type=None,
                   start_color='FFFFFFFF',
                   end_color='FF000000')

highlight.alignment = Alignment(horizontal="center", vertical="center")
highlight.alignment = Alignment(horizontal='center',
                                vertical='center',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0)

customtkinter.set_appearance_mode("dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

app = customtkinter.CTk()
app.geometry("400x580")
app.title("Shifts creator")

wb2 = load_workbook('history.xlsx')
wb = Workbook()
ws = wb.active
wb.add_named_style(highlight)

ws.append(['Sun-Day', 'Sun-Night', 'Mon-Day', 'Mon-Night', 'Tue-Day', 'Tue-Night', 'Wed-Day', 'Wed-Night', 'Thu-Day',
           'Thu-Night', 'Fri-Day', 'Sat-Night'])

val = ""
alyona = ""
alex = ""
ofir = ""
yair = ""
almog = ""
pavel = ""
ran = ""
sahar = ""
hello = ""
maxst = ""
maxfr = ""
shifts = ""
alyona_st = 0
alyona_fr = 0
alex_st = 0
alex_fr = 0
ofir_st = 0
ofir_fr = 0
yair_st = 0
yair_fr = 0
almog_st = 0
almog_fr = 0
ran_st = 0
ran_fr = 0
sahar_st = 0
sahar_fr = 0
pavel_st = 0
pavel_fr = 0


shifts = "shifts: "
i = 0
count = 0
s = 0


def button_callback():
    global val
    val = (" " + val + (optionmenu_1.get()) + " " + (combobox_1.get()) + " ")
    label_2 = customtkinter.CTkLabel(master=frame_1,
                                     text=" " + (optionmenu_1.get()) + " can not work on " + (combobox_1.get()) + " ")
    label_2.pack(pady=1, padx=1)


def check():
    global alyona
    global alex
    global ofir
    global almog
    global yair
    global pavel
    global ran
    global sahar
    global shifts

    # Alyona check
        #DAY
    if "Alyona Sun-Morning" in val:
        alyona = alyona + "ABCD"
    if "Alyona Mon-Morning" in val:
        alyona = alyona + "ABCDG"
    if "Alyona Tue-Morning" in val:
        alyona = alyona + "ABGIJ"
    if "Alyona Wed-Morning" in val:
        alyona = alyona + "BEFIJ"
    if "Alyona Thu-Morning" in val:
        alyona = alyona + "BEFHJ"
    if "Alyona Fri-Morning " in val:
        alyona = alyona + "FHJ"
        # NIGHT
    if "Alyona Sun-Night" in val:
        alyona = alyona + "EFHIJ"
    if "Alyona Mon-Night" in val:
        alyona = alyona + "EFH"
    if "Alyona Tue-Night" in val:
        alyona = alyona + "CDH"
    if "Alyona Wed-Night" in val:
        alyona = alyona + "ACDG"
    if "Alyona Thu-Night" in val:
        alyona = alyona + "ACDGI"
    if "Alyona Sat-Night " in val:
        alyona = alyona + "EGI"
    #alex
    if "Alex Sun-Morning" in val:
        alex = alex + "ABCD"
    if "Alex Mon-Morning" in val:
        alex = alex + "ABCDG"
    if "Alex Tue-Morning" in val:
        alex = alex + "ABGIJ"
    if "Alex Wed-Morning" in val:
        alex = alex + "BEFIJ"
    if "Alex Thu-Morning" in val:
        alex = alex + "BEFHJ"
    if "Alex Fri-Morning " in val:
        alex = alex + "FHJ"
        # NIGHT
    if "Alex Sun-Night" in val:
        alex = alex + "EFHIJ"
    if "Alex Mon-Night" in val:
        alex = alex + "EFH"
    if "Alex Tue-Night" in val:
        alex = alex + "CDH"
    if "Alex Wed-Night" in val:
        alex = alex + "ACDG"
    if "Alex Thu-Night" in val:
        alex = alex + "ACDGI"
    if "Alex Sat-Night " in val:
        alex = alex + "EGI"
    #ofir
    if "Ofir Sun-Morning" in val:
        ofir = ofir + "ABCD"
    if "Ofir Mon-Morning" in val:
        ofir = ofir + "ABCDG"
    if "Ofir Tue-Morning" in val:
        ofir = ofir + "ABGIJ"
    if "Ofir Wed-Morning" in val:
        ofir = ofir + "BEFIJ"
    if "Ofir Thu-Morning" in val:
        ofir = ofir + "BEFHJ"
    if "Ofir Fri-Morning " in val:
        ofir = ofir + "FHJ"
        # NIGHT
    if "Ofir Sun-Night" in val:
        ofir = ofir + "EFHIJ"
    if "Ofir Mon-Night" in val:
        ofir = ofir + "EFH"
    if "Ofir Tue-Night" in val:
        ofir = ofir + "CDH"
    if "Ofir Wed-Night" in val:
        ofir = ofir + "ACDG"
    if "Ofir Thu-Night" in val:
        ofir = ofir + "ACDGI"
    if "Ofir Sat-Night " in val:
        ofir = ofir + "EGI"
    #yair
    if "Yair Sun-Morning" in val:
        yair = yair + "ABCD"
    if "Yair Mon-Morning" in val:
        yair = yair + "ABCDG"
    if "Yair Tue-Morning" in val:
        yair = yair + "ABGIJ"
    if "Yair Wed-Morning" in val:
        yair = yair + "BEFIJ"
    if "Yair Thu-Morning" in val:
        yair = yair + "BEFHJ"
    if "Yair Fri-Morning " in val:
        yair = yair + "FHJ"
        # NIGHT
    if "Yair Sun-Night" in val:
        yair = yair + "EFHIJ"
    if "Yair Mon-Night" in val:
        yair = yair + "EFH"
    if "Yair Tue-Night" in val:
        yair = yair + "CDH"
    if "Yair Wed-Night" in val:
        yair = yair + "ACDG"
    if "Yair Thu-Night" in val:
        yair = yair + "ACDGI"
    if "Yair Sat-Night " in val:
        yair = yair + "EGI"
    #almog
    if "Almog Sun-Morning" in val:
        almog = almog + "ABCD"
    if "Almog Mon-Morning" in val:
        almog = almog + "ABCDG"
    if "Almog Tue-Morning" in val:
        almog = almog + "ABGIJ"
    if "Almog Wed-Morning" in val:
        almog = almog + "BEFIJ"
    if "Almog Thu-Morning" in val:
        almog = almog + "BEFHJ"
    if "Almog Fri-Morning " in val:
        almog = almog + "FHJ"
        # NIGHT
    if "Almog Sun-Night" in val:
        almog = almog + "EFHIJ"
    if "Almog Mon-Night" in val:
        almog = almog + "EFH"
    if "Almog Tue-Night" in val:
        almog = almog + "CDH"
    if "Almog Wed-Night" in val:
        almog = almog + "ACDG"
    if "Almog Thu-Night" in val:
        almog = almog + "ACDGI"
    if "Almog Sat-Night " in val:
        almog = almog + "EGI"

    if "Pavel  Sun-Morning" in val:
        pavel = pavel + "ABCD"
    if "Pavel  Mon-Morning" in val:
        pavel = pavel + "ABCDG"
    if "Pavel  Tue-Morning" in val:
        pavel = pavel + "ABGIJ"
    if "Pavel  Wed-Morning" in val:
        pavel = pavel + "BEFIJ"
    if "Pavel  Thu-Morning" in val:
        pavel = pavel + "BEFHJ"
    if "Pavel  Fri-Morning " in val:
        pavel = pavel + "FHJ"
        # NIGHT
    if "Pavel  Sun-Night" in val:
        pavel = pavel + "EFHIJ"
    if "Pavel  Mon-Night" in val:
        pavel = pavel + "EFH"
    if "Pavel  Tue-Night" in val:
        pavel = pavel + "CDH"
    if "Pavel  Wed-Night" in val:
        pavel = pavel + "ACDG"
    if "Pavel  Thu-Night" in val:
        pavel = pavel + "ACDGI"
    if "Pavel  Sat-Night " in val:
        pavel = pavel + "EGI"

    if "Ran Sun-Morning" in val:
        ran = ran + "ABCD"
    if "Ran Mon-Morning" in val:
        ran = ran + "ABCDG"
    if "Ran Tue-Morning" in val:
        ran = ran + "ABGIJ"
    if "Ran Wed-Morning" in val:
        ran = ran + "BEFIJ"
    if "Ran Thu-Morning" in val:
        ran = ran + "BEFHJ"
    if "Ran Fri-Morning " in val:
        ran = ran + "FHJ"
        # NIGHT
    if "Ran Sun-Night" in val:
        ran = ran + "EFHIJ"
    if "Ran Mon-Night" in val:
        ran = ran + "EFH"
    if "Ran Tue-Night" in val:
        ran = ran + "CDH"
    if "Ran Wed-Night" in val:
        ran = ran + "ACDG"
    if "Ran Thu-Night" in val:
        ran = ran + "ACDGI"
    if "Ran Sat-Night " in val:
        ran = ran + "EGI"

    if "Sahar Sun-Morning" in val:
        sahar = sahar + "ABCD"
    if "Sahar Mon-Morning" in val:
        sahar = sahar + "ABCDG"
    if "Sahar Tue-Morning" in val:
        sahar = sahar + "ABGIJ"
    if "Sahar Wed-Morning" in val:
        sahar = sahar + "BEFIJ"
    if "Sahar Thu-Morning" in val:
        sahar = sahar + "BEFHJ"
    if "Sahar Fri-Morning " in val:
        sahar = sahar + "FHJ"
        # NIGHT
    if "Sahar Sun-Night" in val:
        sahar = sahar + "EFHIJ"
    if "Sahar Mon-Night" in val:
        sahar = sahar + "EFH"
    if "Sahar Tue-Night" in val:
        sahar = sahar + "CDH"
    if "Sahar Wed-Night" in val:
        sahar = sahar + "ACDG"
    if "Sahar Thu-Night" in val:
        sahar = sahar + "ACDGI"
    if "Sahar Sat-Night " in val:
        sahar = sahar + "EGI"

     #check who did sat las week and block sanday morning and sat night
    sheets_list = wb2.sheetnames
    df = pd.read_excel('history.xlsx', sheet_name=sheets_list[-1])
    st = (df['SaturDay'])
    if 'alyona' in st.values:
        alyona = alyona + "ABCDEGI"
    if 'alex' in st.values:
        alex = alex + "ABCDEGI"
    if 'ofir' in st.values:
        ofir = ofir + "ABCD"
    if 'yair' in st.values:
        yair = yair + "ABCDEGI"
    if 'almog' in st.values:
        almog = almog + "ABCDEGI"
    if 'ran' in st.values:
        ran = ran + "ABCDEGI"
    if 'sahar' in st.values:
        sahar = sahar + "ABCDEGI"
    if 'pavel' in st.values:
        pavel = pavel + "ABCDEGI"



def button_callback2():
    global val
    global hello
    global s
    global alyona
    global alex
    global ofir
    global almog
    global yair
    global pavel
    global ran
    global sahar
    # after get from input which shifts each ops can not work, transform the data to format 'X = (can not work) ABC..'
    check()

    # shfits lists
    A = ""
    B = ""
    C = ""
    D = ""
    E = ""
    F = ""
    G = ""
    H = ""
    I = ""
    J = ""
    # check for each Ops which shifts can not work
    if not "A" in alyona:
        A = A + "alyona "
    if not "A" in alex:
        A = A + "alex "
    if not "A" in ofir:
        A = A + "ofir "
    if not "A" in yair:
        A = A + "yair "
    if not "A" in almog:
        A = A + "almog "
    if not "A" in pavel:
        A = A + "pavel "
    if not "A" in sahar:
        A = A + "sahar "
    if not "A" in ran:
        A = A + "ran "
    if not "B" in alyona:
        B = B + "alyona "
    if not "B" in alex:
        B = B + "alex "
    if not "B" in ofir:
        B = B + "ofir "
    if not "B" in yair:
        B = B + "yair "
    if not "B" in almog:
        B = B + "almog "
    if not "B" in pavel:
        B = B + "pavel "
    if not "B" in sahar:
        B = B + "sahar "
    if not "B" in ran:
        B = B + "ran "
    if not "C" in alyona:
        C = C + "alyona "
    if not "C" in alex:
        C = C + "alex "
    if not "C" in ofir:
        C = C + "ofir "
    if not "C" in yair:
        C = C + "yair "
    if not "C" in almog:
        C = C + "almog "
    if not "C" in pavel:
        C = C + "pavel "
    if not "C" in sahar:
        C = C + "sahar "
    if not "C" in ran:
        C = C + "ran "
    if not "D" in alyona:
        D = D + "alyona "
    if not "D" in alex:
        D = D + "alex "
    if not "D" in ofir:
        D = D + "ofir "
    if not "D" in yair:
        D = D + "yair "
    if not "D" in almog:
        D = D + "almog "
    if not "D" in pavel:
        D = D + "pavel "
    if not "D" in sahar:
        D = D + "sahar "
    if not "D" in ran:
        D = D + "ran "
    if 'alyona' in E:
        E = E + "alyona "
    if not "E" in alex:
        E = E + "alex "
    if not "E" in ofir:
        E = E + "ofir "
    if not "E" in yair:
        E = E + "yair "
    if not "E" in almog:
        E = E + "almog "
    if not "E" in pavel:
        E = E + "pavel "
    if not "E" in sahar:
        E = E + "sahar "
    if not "E" in ran:
        E = E + "ran "
    if not "F" in alyona:
        F = F + "alyona "
    if not "F" in alex:
        F = F + "alex "
    if not "F" in ofir:
        F = F + "ofir "
    if not "F" in yair:
        F = F + "yair "
    if not "F" in almog:
        F = F + "almog "
    if not "F" in pavel:
        F = F + "pavel "
    if not "F" in sahar:
        F = F + "sahar "
    if not "F" in ran:
        F = F + "ran "
    if not "G" in alyona:
        G = G + "alyona "
    if not "G" in alex:
        G = G + "alex "
    if not "G" in ofir:
        G = G + "ofir "
    if not "G" in yair:
        G = G + "yair "
    if not "G" in almog:
        G = G + "almog "
    if not "G" in pavel:
        G = G + "pavel "
    if not "G" in sahar:
        G = G + "sahar "
    if not "G" in ran:
        G = G + "ran "
    if not "H" in alyona:
        H = H + "alyona "
    if not "H" in alex:
        H = H + "alex "
    if not "H" in ofir:
        H = H + "ofir "
    if not "H" in yair:
        H = H + "yair "
    if not "H" in almog:
        H = H + "almog "
    if not "H" in pavel:
        H = H + "pavel "
    if not "H" in sahar:
        H = H + "sahar "
    if not "H" in ran:
        H = H + "ran "
    if not "I" in alyona:
        I = I + "alyona "
    if not "I" in alex:
        I = I + "alex "
    if not "I" in ofir:
        I = I + "ofir "
    if not "I" in yair:
        I = I + "yair "
    if not "I" in almog:
        I = I + "almog "
    if not "I" in pavel:
        I = I + "pavel "
    if not "I" in sahar:
        I = I + "sahar "
    if not "I" in ran:
        I = I + "ran "
    if not "J" in alyona:
        J = J + "alyona "
    if not "J" in alex:
        J = J + "alex "
    if not "J" in ofir:
        J = J + "ofir "
    if not "J" in yair:
        J = J + "yair "
    if not "J" in almog:
        J = J + "almog "
    if not "J" in pavel:
        J = J + "pavel "
    if not "J" in sahar:
        J = J + "sahar "
    if not "J" in ran:
        J = J + "ran "


        # Convert to list

    A = A.split(" ")
    A.pop()
    B = B.split(" ")
    B.pop()
    C = C.split(" ")
    C.pop()
    D = D.split(" ")
    D.pop()
    E = E.split(" ")
    E.pop()
    F = F.split(" ")
    F.pop()
    G = G.split(" ")
    G.pop()
    H = H.split(" ")
    H.pop()
    I = I.split(" ")
    I.pop()
    J = J.split(" ")
    J.pop()


    get_data()
    global alyona_st
    global alex_st
    global ofir_st
    global yair_st
    global almog_st
    global ran_st
    global pavel_st
    global sahar_st
    global alyona_fr
    global alex_fr
    global ofir_fr
    global yair_fr
    global almog_fr
    global ran_fr
    global pavel_fr
    global sahar_fr
    global maxst
    global maxfr
    global shifts

    fr = [alyona_fr, alex_fr, ofir_fr, almog_fr, pavel_fr, yair_fr, sahar_fr, pavel_fr, ran_fr]
    st = [alyona_st, alex_st, ofir_st, almog_st, pavel_st, yair_st, sahar_st, pavel_st, ran_st]
    print(st)
    print(min(st))
    print(fr)
    print(min(fr))


    hello = "1 2 3 4 5 6 7 8 9 10" #creat string
    hello = hello.split(" ") #list

    print("A:" ,A)
    print("B:" ,B)
    print("C:" ,C)
    print("D:" ,D)
    print("E:" ,E)
    print("F:" ,F)
    print("G:" ,G)
    print("H:" ,H)
    print("I:" ,I)
    print("J:" ,J)
    print("alyona", alyona)
    print("alex", alex)
    print("almog", almog)
    print("ofir", ofir)
    print("yair", yair)
    print("ran", ran)
    print("pavel", pavel)
    print("sahar", sahar)



    for x in range(0, 50):
        s = 0
        if E:
            for i in range(0, 3):
                if (alyona_st <= min(st) + i) and ('alyona' in E) and (not "alyona" in hello):
                    hello[4] = ("alyona")
                    break
                if (alex_st <= min(st) + i) and ('alex' in E) and (not "alex" in hello):
                    hello[4] = ("alex")
                    break
                if (ofir_st <= min(st) + i) and ('ofir' in E) and (not "ofir" in hello):
                    hello[4] = ("ofir")
                    break
                if (yair_st <= min(st) + i) and ('yair' in E) and (not "yair" in hello):
                    hello[4] = ("yair")
                    break
                if (pavel_st <= min(st) + i) and ('pavel' in E) and (not "pavel" in hello):
                    hello[4] = ("pavel")
                    break
                if (sahar_st <= min(st) + i) and ('sahar' in E) and (not "sahar" in hello):
                    hello[4] = ("sahar")
                    break
                if (ran_st <= min(st) + i) and ('ran' in E) and (not "ran" in hello):
                    hello[4] = ("ran")
                    break
                if (almog_st <= min(st) + i) and ('almog' in E) and (not "almog" in hello):
                    hello[4] = ("almog")
                    break
        elif (not E):
            hello[4] = ("NOTE")
            s = s + 1
        if G:
            for i in range(0, 3):
                if (alyona_st <= min(st) + i) and ("alyona" in G) and (not "alyona" in hello):
                    hello[6] = ("alyona")
                    break
                if (alex_st <= min(st) + i) and ("alex" in G) and (not "alex" in hello):
                    hello[6] = ("alex")
                    break
                if (ofir_st <= min(st) + i) and ("ofir" in G) and (not "ofir" in hello):
                    hello[6] = ("ofir")
                    break
                if (yair_st <= min(st) + i) and ("yair" in G) and (not "yair" in hello):
                    hello[6] = ("yair")
                    break
                if (pavel_st <= min(st) + i) and ("pavel" in G) and (not "pavel" in hello):
                    hello[6] = ("pavel")
                    break
                if (sahar_st <= min(st) + i) and ("sahar" in G) and (not "sahar" in hello):
                    hello[6] = ("sahar")
                    break
                if (ran_st <= min(st) + i) and ("ran" in G) and (not "ran" in hello):
                    hello[6] = ("ran")
                    break
                if (almog_st <= min(st) + i) and ("almog" in G) and (not "almog" in hello):
                    hello[6] = ("almog")
                    break
        elif (not G):
            hello[6] = ("NOTG")
            s = s + 1
        if F:
            for i in range(0, 3):
                if (alyona_fr <= min(fr) + i) and ('alyona' in F) and (not "alyona" in hello):
                    hello[5] = ("alyona")
                    break
                if (alex_fr <= min(fr) + i) and ('alex' in F) and (not "alex" in hello):
                    hello[5] = ("alex")
                    break
                if (ofir_fr <= min(fr) + i) and ('ofir' in F) and (not "ofir" in hello):
                    hello[5] = ("ofir")
                    break
                if (yair_fr <= min(fr) + i) and ('yair' in F) and (not "yair" in hello):
                    hello[5] = ("yair")
                    break
                if (pavel_fr <= min(fr) + i) and ('pavel' in F) and (not "pavel" in hello):
                    hello[5] = ("pavel")
                    break
                if (sahar_fr <= min(fr) + i) and ('sahar' in F) and (not "sahar" in hello):
                    hello[5] = ("sahar")
                    break
                if (ran_fr <= min(fr) + i) and ('ran' in F) and (not "ran" in hello):
                    hello[5] = ("ran")
                    break
                if (almog_fr <= min(fr) + i) and ('almog' in F) and (not "almog" in hello):
                    hello[5] = ("almog")
                    break
        elif (not F):
            hello[5] = ("NOTF")
            s = s + 1
        if H:
            for i in range(0, 3):
                if (alyona_fr <= min(fr) + i) and ("alyona" in H) and (not "alyona" in hello):
                    hello[7] = ("alyona")
                    break
                if (alex_fr <= min(fr) + i) and ("alex" in H) and (not "alex" in hello):
                    hello[7] = ("alex")
                    break
                if (ofir_fr <= min(fr) + i) and ("ofir" in H) and (not "ofir" in hello):
                    hello[7] = ("ofir")
                    break
                if (yair_fr <= min(fr) + i) and ("yair" in H) and (not "yair" in hello):
                    hello[7] = ("yair")
                    break
                if (pavel_fr <= min(fr) + i) and ("pavel" in H) and (not "pavel" in hello):
                    hello[7] = ("pavel")
                    break
                if (sahar_fr <= min(fr) + i) and ("sahar" in H) and (not "sahar" in hello):
                    hello[7] = ("sahar")
                    break
                if (ran_fr <= min(fr) + i) and ("ran" in H) and (not "ran" in hello):
                    hello[7] = ("ran")
                    break
                if (almog_fr <= min(fr) + i) and ("almog" in H) and (not "almog" in hello):
                    hello[7] = ("almog")
                    break
        elif not H:
            hello[7] = ("NOT H")
            s = s + 1
        if A:
            for x in range(0, 99):
                hello[0] = (random.choice(A))
                if hello.count(hello[0]) == 1:
                    break
        elif (not A):
            hello[0] = ("NOTA")
            s = s + 1
        if B:
            for x in range(0, 99):
                hello[1] = (random.choice(B))
                if hello.count(hello[1]) == 1:
                    break
        elif (not B):
            hello[1] = ("B")
            s = s + 1
        if C:
            for x in range(0, 99):
                hello[2] = (random.choice(C))
                if hello.count(hello[2]) == 2:
                    break
        elif (not C):
            hello[0] = ("C")
            s = s + 2
        if D:
            for x in range(0, 99):
                hello[3] = (random.choice(D))
                if hello.count(hello[3]) == 1:
                    break
        elif (not D):
            hello[3] = ("NOTD")
            s = s + 1
        if (I) and (s>0):
            for x in range(0, 99):
                hello[8] = (random.choice(I))
                if hello.count(hello[8]) == 1:
                    break
        elif (not I):
            hello[8] = ("NOTI")
            s = s + 1
        if (J) and (s>1):
            for x in range(0, 99):
                hello[9] = (random.choice(I))
                if hello.count(hello[9]) == 1:
                    break
        elif (not J):
            hello[9] = ("NOTJ")

        if (hello.count('alyona') == 1) and (hello.count('alex') == 1) and (
                hello.count('ofir') == 1) and (hello.count('yair') == 1) and (hello.count('almog') == 1) and (
                hello.count('pavel') == 1) and (hello.count('ran') == 1) and (hello.count('sahar') == 1) and (hello.count('C') == 0) and (hello.count('G') == 0):
            break



    his()
    design()
    # df = pd.read_excel(r'Path where the Excel file is stored\shifts.xlsx')
    # print(df)
    absolutePath = Path('../SC/shifts.xlsx').resolve()
    os.system(f'start excel.exe "{absolutePath}"')
    # print(st, fr)
    # ws.append(
    #     "A:" + hello[0] + " B:" + hello[1] + " C:" + hello[2] + " D:" + hello[3] + " E:" + hello[4] + " F:" + hello[
    #         5] + " G:" + hello[6] + " H:" + hello[7] + " I:" + hello[8] + " J:" + hello[9])

    # for x in range(41, 70):
    #     if (ws.cell(row=1, column=x).value) == None:
    #         ws.append(hello)
    #         wb.save('shifts.xlsx')
    #         break


def design():
    # excel write
    ws['A40'] = ("history")
    ws['A40'].style = highlight
    ws['A2'] = (hello[0])
    ws['A3'] = (hello[1])
    ws['A4'] = (hello[2])
    ws['A5'] = (hello[3])
    ws['C2'] = (hello[0])
    ws['C3'] = (hello[1])
    ws['C4'] = (hello[2])
    ws['C5'] = (hello[3])
    ws['E2'] = (hello[0])
    ws['E3'] = (hello[1])
    ws['F4'] = (hello[2])
    ws['F5'] = (hello[3])
    ws['H2'] = (hello[0])
    ws['G3'] = (hello[1])
    ws['H4'] = (hello[2])
    ws['H5'] = (hello[3])
    ws['J2'] = (hello[0])
    ws['J4'] = (hello[2])
    ws['J5'] = (hello[3])
    ws['B6'] = (hello[4])
    ws['B7'] = (hello[5])
    ws['C8'] = (hello[6])
    ws['B9'] = (hello[7])
    ws['B10'] = (hello[8])
    ws['B11'] = (hello[9])
    ws['D6'] = (hello[4])
    ws['D7'] = (hello[5])
    ws['E8'] = (hello[6])
    ws['D9'] = (hello[7])
    ws['E10'] = (hello[8])
    ws['E11'] = (hello[9])
    ws['G6'] = (hello[4])
    ws['G7'] = (hello[5])
    ws['H8'] = (hello[6])
    ws['F9'] = (hello[7])
    ws['G10'] = (hello[8])
    ws['G11'] = (hello[9])
    ws['I3'] = (hello[1])
    ws['I6'] = (hello[4])
    ws['I7'] = (hello[5])
    ws['J8'] = (hello[6])
    ws['I9'] = (hello[7])
    ws['J10'] = (hello[8])
    ws['I11'] = (hello[9])
    ws['L6'] = (hello[4])
    ws['K7'] = (hello[5])
    ws['L8'] = (hello[6])
    ws['K9'] = (hello[7])
    ws['L10'] = (hello[8])
    ws['K11'] = (hello[9])
    # excel design
    ws['A1'].style = highlight
    ws['A2'].style = highlight
    ws['A3'].style = highlight
    ws['A4'].style = highlight
    ws['A5'].style = highlight
    ws['A6'].style = highlight
    ws['A7'].style = highlight
    ws['A8'].style = highlight
    ws['A9'].style = highlight
    ws['A10'].style = highlight
    ws['A11'].style = highlight
    ws['B1'].style = highlight
    ws['B2'].style = highlight
    ws['B3'].style = highlight
    ws['B4'].style = highlight
    ws['B5'].style = highlight
    ws['B6'].style = highlight
    ws['B7'].style = highlight
    ws['B8'].style = highlight
    ws['B9'].style = highlight
    ws['B10'].style = highlight
    ws['B11'].style = highlight
    ws['C1'].style = highlight
    ws['C2'].style = highlight
    ws['C3'].style = highlight
    ws['C4'].style = highlight
    ws['C5'].style = highlight
    ws['C6'].style = highlight
    ws['C7'].style = highlight
    ws['C8'].style = highlight
    ws['C9'].style = highlight
    ws['C10'].style = highlight
    ws['C11'].style = highlight
    ws['D1'].style = highlight
    ws['D2'].style = highlight
    ws['D3'].style = highlight
    ws['D4'].style = highlight
    ws['D5'].style = highlight
    ws['D6'].style = highlight
    ws['D7'].style = highlight
    ws['D8'].style = highlight
    ws['D9'].style = highlight
    ws['D10'].style = highlight
    ws['D11'].style = highlight
    ws['E1'].style = highlight
    ws['E2'].style = highlight
    ws['E3'].style = highlight
    ws['E4'].style = highlight
    ws['E5'].style = highlight
    ws['E6'].style = highlight
    ws['E7'].style = highlight
    ws['E8'].style = highlight
    ws['E9'].style = highlight
    ws['E10'].style = highlight
    ws['E11'].style = highlight
    ws['F1'].style = highlight
    ws['F2'].style = highlight
    ws['F3'].style = highlight
    ws['F4'].style = highlight
    ws['F5'].style = highlight
    ws['F6'].style = highlight
    ws['F7'].style = highlight
    ws['F8'].style = highlight
    ws['F9'].style = highlight
    ws['F10'].style = highlight
    ws['F11'].style = highlight
    ws['G1'].style = highlight
    ws['G2'].style = highlight
    ws['G3'].style = highlight
    ws['G4'].style = highlight
    ws['G5'].style = highlight
    ws['G6'].style = highlight
    ws['G7'].style = highlight
    ws['G8'].style = highlight
    ws['G9'].style = highlight
    ws['G10'].style = highlight
    ws['G11'].style = highlight
    ws['H1'].style = highlight
    ws['H2'].style = highlight
    ws['H3'].style = highlight
    ws['H4'].style = highlight
    ws['H5'].style = highlight
    ws['H6'].style = highlight
    ws['H7'].style = highlight
    ws['H8'].style = highlight
    ws['H9'].style = highlight
    ws['H10'].style = highlight
    ws['H11'].style = highlight
    ws['I1'].style = highlight
    ws['I2'].style = highlight
    ws['I3'].style = highlight
    ws['I4'].style = highlight
    ws['I5'].style = highlight
    ws['I6'].style = highlight
    ws['I7'].style = highlight
    ws['I8'].style = highlight
    ws['I9'].style = highlight
    ws['I10'].style = highlight
    ws['I11'].style = highlight
    ws['J1'].style = highlight
    ws['J2'].style = highlight
    ws['J3'].style = highlight
    ws['J4'].style = highlight
    ws['J5'].style = highlight
    ws['J6'].style = highlight
    ws['J7'].style = highlight
    ws['J8'].style = highlight
    ws['J9'].style = highlight
    ws['J10'].style = highlight
    ws['J11'].style = highlight
    ws['K1'].style = highlight
    ws['K2'].style = highlight
    ws['K3'].style = highlight
    ws['K4'].style = highlight
    ws['K5'].style = highlight
    ws['K6'].style = highlight
    ws['K7'].style = highlight
    ws['K8'].style = highlight
    ws['K9'].style = highlight
    ws['K10'].style = highlight
    ws['K11'].style = highlight
    ws['L1'].style = highlight
    ws['L2'].style = highlight
    ws['L3'].style = highlight
    ws['L4'].style = highlight
    ws['L5'].style = highlight
    ws['L6'].style = highlight
    ws['L7'].style = highlight
    ws['L8'].style = highlight
    ws['L9'].style = highlight
    ws['L10'].style = highlight
    ws['L11'].style = highlight
    ws['A1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['B1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['C1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['D1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['E1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['F1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['G1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['H1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['I1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['J1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['K1'].fill = GradientFill(stop=("0099CCFF", "0099CCFF"))
    ws['L1'].fill = GradientFill(stop=("B3FFB3", "B3FFB3"))
    ws['B2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['B11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['D11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['F11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['H11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['J11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L2'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L3'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L4'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L5'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L6'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L7'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L8'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L9'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L10'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    ws['L11'].fill = GradientFill(stop=("CEFACF", "CEFACF"))
    wb.save('shifts.xlsx')


def button_clear():
    global val
    global shifts
    global alyona
    global alex
    global ofir
    global almog
    global yair
    global pavel
    global ran
    global sahar
    global shifts
    global alyona_st
    global alex_st
    global ofir_st
    global yair_st
    global almog_st
    global ran_st
    global pavel_st
    global sahar_st
    global alyona_fr
    global alex_fr
    global ofir_fr
    global yair_fr
    global almog_fr
    global ran_fr
    global pavel_fr
    global sahar_fr
    global hello
    global maxst
    global maxfr
    global A
    global B
    global C
    global D
    global E
    global F
    global G
    global H
    global I
    global j

    A = ""
    B = ""
    C = ""
    D = ""
    E = ""
    F = ""
    G = ""
    H = ""
    I = ""
    J = ""
    val = ""
    alyona = ""
    alex = ""
    ofir = ""
    yair = ""
    almog = ""
    pavel = ""
    ran = ""
    sahar = ""
    hello = ""
    maxst = ""
    maxfr = ""
    shifts = ""
    alyona_st = 0
    alyona_fr = 0
    alex_st = 0
    alex_fr = 0
    ofir_st = 0
    ofir_fr = 0
    yair_st = 0
    yair_fr = 0
    almog_st = 0
    almog_fr = 0
    ran_st = 0
    ran_fr = 0
    sahar_st = 0
    sahar_fr = 0
    pavel_st = 0
    pavel_fr = 0

    label_5 = customtkinter.CTkLabel(master=frame_1, text=("You have clear all request"))
    label_5.pack(pady=1, padx=1)


def his():
    ex = ["SaturDay", "FriDay"]
    data = [[hello[4], hello[5]], [hello[6], hello[7]], [hello[8], hello[9]]]
    df = pd.DataFrame(data, columns=ex)
    print(df)
    append_df_to_excel('history.xlsx', df, sheet_name='his', index=True, startrow=0)



def get_data():
    global alyona_st
    global alex_st
    global ofir_st
    global yair_st
    global almog_st
    global ran_st
    global pavel_st
    global sahar_st
    global alyona_fr
    global alex_fr
    global ofir_fr
    global yair_fr
    global almog_fr
    global ran_fr
    global pavel_fr
    global sahar_fr
    sheets_list = wb2.sheetnames
    # range (i, j) and if j < 30 delete first item
    for i in range(0, 100):
        if 0 <= i < len(sheets_list):
            df = pd.read_excel('history.xlsx', sheet_name=sheets_list[i])
            st = (df['SaturDay'])
            fr = (df['FriDay'])
            if 'alyona' in st.values:
                alyona_st = alyona_st + 1
            if 'alyona' in fr.values:
                alyona_fr = alyona_fr + 1
            if 'alex' in st.values:
                alex_st = alex_st + 1
            if 'alex' in fr.values:
                alex_fr = alex_fr + 1
            if 'ofir' in st.values:
                ofir_st = ofir_st + 1
            if 'ofir' in fr.values:
                ofir_fr = ofir_fr + 1
            if 'yair' in st.values:
                yair_st = yair_st + 1
            if 'yair' in fr.values:
                yair_fr = yair_fr + 1
            if 'almog' in st.values:
                almog_st = almog_st + 1
            if 'almog' in fr.values:
                almog_fr = almog_fr + 1
            if 'ran' in st.values:
                ran_st = ran_st + 1
            if 'ran' in fr.values:
                ran_fr = ran_fr + 1
            if 'sahar' in st.values:
                sahar_st = sahar_st + 1
            if 'sahar' in fr.values:
                sahar_fr = sahar_fr + 1
            if 'pavel' in st.values:
                pavel_st = pavel_st + 1
            if 'pavel' in fr.values:
                pavel_fr = pavel_fr + 1








def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='new')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


frame_1 = customtkinter.CTkFrame(master=app)
frame_1.pack(pady=60, padx=60, fill="both", expand=True)

label_1 = customtkinter.CTkLabel(master=frame_1, text="Shifts Creator", justify=tkinter.LEFT)
label_1.pack(pady=12, padx=10)

# set the text


# entry_1 = customtkinter.CTkEntry(master=frame_1, placeholder_text="CTkEntry")
# entry_1.pack(pady=12, padx=10)

optionmenu_1 = customtkinter.CTkOptionMenu(frame_1,
                                           values=["Alyona", "Alex", "Ofir", "Yair", "Almog", "Pavel", "Ran", "Sahar"])
optionmenu_1.pack(pady=12, padx=10)
optionmenu_1.set("Alyona")

combobox_1 = customtkinter.CTkComboBox(frame_1, values=["Sun-Morning", "Mon-Morning", "Tue-Morning", "Wed-Morning",
                                                        "Thu-Morning", "Fri-Morning", "Sun-Night", "Mon-Night",
                                                        "Tue-Night", "Wed-Night", "Thu-Night", "Sat-Night"])
combobox_1.pack(pady=12, padx=10)
combobox_1.set("Sun-Morning")

button_1 = customtkinter.CTkButton(master=frame_1, text="save", command=button_callback)
button_1.pack(pady=12, padx=10)

button_2 = customtkinter.CTkButton(master=frame_1, text="finished", command=button_callback2)
button_2.pack(pady=12, padx=10)

button_3 = customtkinter.CTkButton(master=frame_1, text="Clear all", command=button_clear)
button_3.pack(pady=10, padx=8)

label_3 = customtkinter.CTkLabel(master=frame_1, text="Shifts blocked:", justify=tkinter.LEFT)
label_3.pack(pady=5, padx=3)

app.mainloop()
